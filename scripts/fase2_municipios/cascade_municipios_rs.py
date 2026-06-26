#!/usr/bin/env python3
"""5-tier cascade pipeline for RS municipality resource discovery.

Finds the stable index/listing page for concursos and processos seletivos
in each RS municipality. Does NOT extract individual editals.

Architecture (spend expensive tools only when cheap ones fail):
    Tier 0 — Site oficial: find/confirm the prefeitura's base domain.
    Tier 1 — Free link discovery: HTML menus, anchors, sitemap, transparência.
    Tier 2 — Grounded search: Gemini + Google Search (only if Tier 1 incomplete).
    Tier 3 — Gemini verifier/selector: classifies candidates with discrete
             decisions (indice_oficial, detalle_rechazado, licitacao_rechazada,
             etc.) and picks best among valid ones (ai_pick_best).
    Tier 4 — Playwright navigation agent: directed menu navigation as last resort.

No numeric scorers. No magic constants. Discrete decisions + AI judgment.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
import traceback
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin, urlparse, unquote

import requests

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_MUNICIPIOS_URL = "https://dados.tce.rs.gov.br/dados/auxiliar/municipios.csv"
UF_SIGLA = "RS"
UF_NOME = "Rio Grande do Sul"
GEMINI_BASE_URL = "https://generativelanguage.googleapis.com/v1beta"

BAD_HOSTS = [
    "facebook.", "instagram.", "youtube.", "twitter.", "x.com",
    "linkedin.", "acheconcursos.", "pciconcursos.", "qconcursos.",
    "google.", "bing.", "duckduckgo.",
]

BUCKET_KEYWORDS = {
    "concursos": [
        "concurso", "concursos", "concursos publicos", "concurso publico",
    ],
    "processos": [
        "processo seletivo", "processos seletivos", "pss",
        "processos seletivos simplificados", "selecao publica",
        "selecoes publicas", "seletivo simplificado",
    ],
}

CONTAINER_KEYWORDS = [
    "editais", "edital", "publicacoes", "publicacao", "documentos",
    "transparencia", "contratacao", "oportunidades", "mural",
    "servicos", "portal",
]


# ---------------------------------------------------------------------------
# Text utilities
# ---------------------------------------------------------------------------
def norm(text: str) -> str:
    t = unicodedata.normalize("NFKD", text or "")
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = re.sub(r"[^a-z0-9\s]+", " ", t.lower())
    return re.sub(r"\s+", " ", t).strip()


def slugify(name: str) -> str:
    n = norm(name)
    n = re.sub(r"\b(da|de|do|das|dos)\b", "", n)
    return re.sub(r"[^a-z0-9]+", "", n)


def clean_url(url: str) -> str:
    url = (url or "").strip()
    if url and "://" not in url:
        url = "http://" + url
    return url.rstrip("/") if url else ""


# ---------------------------------------------------------------------------
# Page dataclass
# ---------------------------------------------------------------------------
@dataclass
class Page:
    url: str
    status: int = 0
    title: str = ""
    text: str = ""
    links: list[tuple[str, str]] = field(default_factory=list)
    error: str = ""

    @property
    def ok(self) -> bool:
        return 200 <= self.status < 400 and not self.error


# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------
def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.5",
    })
    return s


def fetch_page(session: requests.Session, url: str, timeout: int = 15) -> Page:
    url = clean_url(url)
    if not url:
        return Page(url="", error="empty_url")
    try:
        resp = session.get(url, timeout=timeout, allow_redirects=True)
        content_type = resp.headers.get("content-type", "")
        if "text/html" not in content_type and "text/plain" not in content_type:
            return Page(url=resp.url, status=resp.status_code, error="not_html")
        html_text = resp.text
        title = ""
        m = re.search(r"<title[^>]*>(.*?)</title>", html_text, re.I | re.S)
        if m:
            title = re.sub(r"\s+", " ", m.group(1)).strip()
        links = extract_links(resp.url, html_text)
        body_text = extract_text(html_text)
        return Page(url=resp.url, status=resp.status_code, title=title,
                    text=body_text, links=links)
    except Exception as e:
        return Page(url=url, error=str(e)[:200])


def extract_links(base_url: str, html: str) -> list[tuple[str, str]]:
    results = []
    seen = set()
    for m in re.finditer(r'<a\s[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', html, re.I | re.S):
        href_raw, link_text = m.group(1), m.group(2)
        link_text = re.sub(r"<[^>]+>", "", link_text).strip()
        link_text = re.sub(r"\s+", " ", link_text)
        try:
            href = urljoin(base_url, href_raw)
        except Exception:
            continue
        if href not in seen and href.startswith("http"):
            seen.add(href)
            results.append((href, link_text))
    return results


def extract_text(html: str) -> str:
    text = re.sub(r"<script[^>]*>.*?</script>", " ", html, flags=re.I | re.S)
    text = re.sub(r"<style[^>]*>.*?</style>", " ", text, flags=re.I | re.S)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"&[a-zA-Z]+;", " ", text)
    text = re.sub(r"&#\d+;", " ", text)
    return re.sub(r"\s+", " ", text).strip()


# ---------------------------------------------------------------------------
# Basic filters (cheap, deterministic, obvious rejections only)
# ---------------------------------------------------------------------------
SOFT_404_PATTERNS = [
    "pagina nao encontrada", "nao encontramos", "erro 404", "error 404",
    "not found", "pagina inexistente", "conteudo nao encontrado",
]


def is_soft_404(page: Page) -> bool:
    if not page.ok:
        return True
    blob = norm(page.title + " " + page.text[:500])
    return any(p in blob for p in SOFT_404_PATTERNS)


def is_broad_landing(url: str) -> bool:
    path = (urlparse(clean_url(url)).path or "/").strip("/").lower()
    return path in {"", "web", "home", "inicio", "index.php", "index.html",
                    "portal", "site"}


def is_pdf_or_file(url: str) -> bool:
    path = urlparse(url).path.lower()
    return any(path.endswith(ext) for ext in [".pdf", ".doc", ".docx", ".xls", ".xlsx"])


# ---------------------------------------------------------------------------
# Candidate: a URL with metadata
# ---------------------------------------------------------------------------
@dataclass
class Candidate:
    url: str
    source: str  # "menu_link", "container_link", "grounding", "playwright"
    menu_text: str = ""
    page: Page | None = None
    fetchable: bool = True
    content_preview: str = ""


# ---------------------------------------------------------------------------
# TIER 0: Find/confirm site base
# ---------------------------------------------------------------------------
def domain_candidates(municipio: str) -> list[str]:
    # Only the safe, full-name slugs (no collision-prone heuristics like
    # first-word or pm+initials). When these miss, Tier 2 grounded search
    # discovers the real domain — see tier2_find_site_grounded.
    slugs = []
    s1 = slugify(municipio)
    s2 = re.sub(r"[^a-z0-9]+", "", norm(municipio))
    for s in [s1, s2]:
        if s and s not in slugs:
            slugs.append(s)
    urls = []
    for slug in slugs:
        for prefix in ["www.", "", "pm"]:
            host = f"{prefix}{slug}.rs.gov.br"
            urls.extend([f"https://{host}/", f"http://{host}/"])
        urls.append(f"https://{slug}.atende.net/")
    seen = set()
    return [u for u in urls if u not in seen and not seen.add(u)]


def score_site_page(page: Page, municipio: str) -> int:
    """How strongly a page looks like the official prefeitura homepage.

    Used both by the free slug discovery (Tier 0) and the grounded domain
    discovery fallback, so the validation bar is identical regardless of how
    the URL was found.
    """
    blob = norm(page.title + " " + page.text[:2000])
    muni_norm = norm(municipio)
    score = 0
    if muni_norm in blob:
        score += 10
    if "prefeitura" in blob:
        score += 5
    if ".rs.gov.br" in page.url:
        score += 3
    if ".atende.net" in page.url:
        score += 2
    return score


def tier0_find_site(session: requests.Session, municipio: str,
                    timeout: int = 15) -> Page | None:
    candidates = domain_candidates(municipio)
    best = None
    best_score = -1
    for url in candidates:
        page = fetch_page(session, url, timeout)
        if not page.ok:
            continue
        score = score_site_page(page, municipio)
        if score > best_score:
            best_score = score
            best = page
    if best and best_score >= 5:
        migrated = _check_migration(session, best, timeout)
        if migrated:
            return migrated
        return best
    return best if best and best_score > 0 else None


def _check_migration(session: requests.Session, page: Page,
                     timeout: int = 15) -> Page | None:
    migration_patterns = [
        r"novo\s+site", r"novo\s+endereco", r"novo\s+portal",
        r"mudou\s+para", r"acesse\s+o\s+novo",
    ]
    blob = norm(page.text[:3000])
    if not any(re.search(p, blob) for p in migration_patterns):
        return None
    for href, text in page.links:
        text_n = norm(text)
        if any(re.search(p, text_n) for p in migration_patterns) or "novo site" in text_n:
            new_page = fetch_page(session, href, timeout)
            if new_page.ok and not is_soft_404(new_page):
                return new_page
    return None


# ---------------------------------------------------------------------------
# TIER 1: Free link discovery
# ---------------------------------------------------------------------------
def tier1_collect_candidates(session: requests.Session, home: Page,
                             municipio: str, timeout: int = 15) -> list[Candidate]:
    """Scan home page links and one level of container pages for relevant URLs."""
    candidates = []
    seen_urls: set[str] = set()
    all_keywords = []
    for kws in BUCKET_KEYWORDS.values():
        all_keywords.extend(kws)

    # Direct links from home page
    for href, link_text in home.links:
        host = urlparse(href).netloc.lower()
        if any(bad in host for bad in BAD_HOSTS):
            continue
        if is_pdf_or_file(href) or is_broad_landing(href):
            continue
        text_n = norm(link_text)
        href_n = norm(unquote(urlparse(href).path))
        if any(kw in text_n or kw in href_n for kw in all_keywords):
            if href not in seen_urls:
                seen_urls.add(href)
                candidates.append(Candidate(
                    url=href, source="menu_link", menu_text=link_text,
                ))

    # One level deep: follow container-like links
    # Skip links already captured as bucket candidates
    bucket_hrefs = {c.url for c in candidates}
    container_urls = []
    for href, link_text in home.links:
        if href in bucket_hrefs or href in seen_urls:
            continue
        text_n = norm(link_text)
        href_n = norm(unquote(urlparse(href).path))
        if any(kw in text_n or kw in href_n for kw in CONTAINER_KEYWORDS):
            if not is_broad_landing(href) and not is_pdf_or_file(href):
                seen_urls.add(href)
                container_urls.append((href, link_text))

    for container_href, container_text in container_urls[:4]:
        container_page = fetch_page(session, container_href, min(timeout, 10))
        if not container_page.ok or is_soft_404(container_page):
            continue
        for href, link_text in container_page.links:
            host = urlparse(href).netloc.lower()
            if any(bad in host for bad in BAD_HOSTS):
                continue
            if is_pdf_or_file(href) or is_broad_landing(href):
                continue
            text_n = norm(link_text)
            href_n = norm(unquote(urlparse(href).path))
            if any(kw in text_n or kw in href_n for kw in all_keywords):
                if href not in seen_urls:
                    seen_urls.add(href)
                    candidates.append(Candidate(
                        url=href, source="container_link",
                        menu_text=f"{container_text} > {link_text}",
                    ))

    # Fetch each candidate page to get content for Tier 3
    for c in candidates:
        page = fetch_page(session, c.url, min(timeout, 10))
        if page.ok and not is_soft_404(page):
            c.page = page
            c.content_preview = page.text[:1200]
            c.fetchable = True
        else:
            c.fetchable = False

    # Drill-down: a bucket parent page (e.g. /concurso) often links to more
    # specific sub-indexes (e.g. /concurso/categoria/25/concurso). Also follows
    # same-level siblings (same path depth, different leaf) so that a concursos
    # page can lead to the processos page beside it.
    drill: list[Candidate] = []
    for c in [c for c in candidates if c.fetchable and c.page]:
        parent_path = urlparse(c.url).path.rstrip("/")
        parent_host = urlparse(c.url).netloc.lower()
        parent_parent = "/".join(parent_path.split("/")[:-1]) if "/" in parent_path.lstrip("/") else ""
        for href, link_text in c.page.links:
            if href in seen_urls or len(drill) >= 12:
                continue
            pu = urlparse(href)
            if pu.netloc.lower() != parent_host:
                continue
            child_path = pu.path.rstrip("/")
            is_child = child_path.startswith(parent_path + "/")
            is_sibling = (parent_parent
                          and child_path.startswith(parent_parent + "/")
                          and child_path != parent_path
                          and child_path.count("/") == parent_path.count("/"))
            if not is_child and not is_sibling:
                continue
            if is_pdf_or_file(href) or is_broad_landing(href):
                continue
            text_n = norm(link_text)
            href_n = norm(unquote(pu.path))
            if any(kw in text_n or kw in href_n for kw in all_keywords):
                seen_urls.add(href)
                drill.append(Candidate(
                    url=href, source="drilldown",
                    menu_text=f"{c.menu_text} > {link_text}",
                ))
    for c in drill:
        page = fetch_page(session, c.url, min(timeout, 10))
        if page.ok and not is_soft_404(page):
            c.page = page
            c.content_preview = page.text[:1200]
            c.fetchable = True
        else:
            c.fetchable = False
    candidates.extend(drill)

    # Parameter normalization: if a candidate has ano=YYYY, add ano=0 variant
    # (all years) so Tier 3 can pick the canonical unfiltered view.
    param_variants: list[Candidate] = []
    for c in [c for c in candidates if c.fetchable]:
        if "ano=" in c.url and not re.search(r"[?&]ano=0(?:&|$)", c.url):
            variant_url = re.sub(r"([?&]ano=)\d{4}", r"\g<1>0", c.url)
            if variant_url not in seen_urls:
                seen_urls.add(variant_url)
                param_variants.append(Candidate(
                    url=variant_url, source="param_variant",
                    menu_text=f"{c.menu_text} (all years)",
                ))
    for c in param_variants:
        page = fetch_page(session, c.url, min(timeout, 10))
        if page.ok and not is_soft_404(page):
            c.page = page
            c.content_preview = page.text[:1200]
            c.fetchable = True
        else:
            c.fetchable = False
    candidates.extend(param_variants)

    return candidates


# ---------------------------------------------------------------------------
# TIER 2: Gemini grounded search
# ---------------------------------------------------------------------------
def gemini_api_key() -> str:
    return os.environ.get("GEMINI_API_KEY", "")


def gemini_post(session: requests.Session, model: str, payload: dict,
                timeout: int = 90) -> dict:
    key = gemini_api_key()
    if not key:
        raise RuntimeError("missing GEMINI_API_KEY")
    url = f"{GEMINI_BASE_URL}/models/{model}:generateContent?key={key}"
    for attempt in range(2):
        try:
            resp = session.post(url, json=payload, timeout=timeout)
            if resp.status_code == 429:
                time.sleep(4 * (attempt + 1))
                continue
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            print(f"      gemini attempt {attempt+1} failed: {e}", flush=True)
            if attempt == 1:
                raise
            time.sleep(4)
    return {}


def tier2_grounded_search(session: requests.Session, model: str,
                          municipio: str, site_hint: str,
                          timeout: int = 15) -> list[Candidate]:
    hint = f"O site oficial e: {site_hint}. " if site_hint else ""
    prompt = (
        f"Voce e um investigador de sites oficiais de prefeituras do {UF_NOME} ({UF_SIGLA}), Brasil. "
        f"{hint}"
        f"Encontre no Google as URLs OFICIAIS e ESTAVEIS da prefeitura de "
        f"{municipio} ({UF_NOME}, {UF_SIGLA}, Brasil) para concursos publicos "
        f"e processos seletivos.\n"
        "REGRAS:\n"
        f"- Busque SEMPRE incluindo '{UF_NOME}' ou '{UF_SIGLA}'.\n"
        "- Prefira dominio oficial (.rs.gov.br ou .atende.net).\n"
        "- Queremos a PAGINA INDICE/LISTAGEM (onde se listam varios editais), "
        "NAO um edital individual nem um PDF.\n"
        "- NAO use licitacoes, pregao, compras, chamamento publico.\n"
        "Liste com a URL completa: site oficial; pagina de concursos; pagina de PSS."
    )
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {"temperature": 1.0, "maxOutputTokens": 2048},
    }
    data = gemini_post(session, model, payload, timeout=90)

    cand = (data.get("candidates") or [{}])[0]
    text_response = "\n".join(
        p.get("text", "") for p in (cand.get("content", {}) or {}).get("parts", [])
        if isinstance(p, dict)
    )
    chunks = (cand.get("groundingMetadata", {}) or {}).get("groundingChunks", []) or []

    candidates = []
    seen: set[str] = set()

    # Host filter: only accept URLs from the official site's domain
    hint_host = ""
    if site_hint:
        hint_host = urlparse(site_hint).netloc.lower().lstrip("www.")

    def _t2_host_ok(h: str) -> bool:
        if not hint_host:
            return True
        h = h.lower().lstrip("www.")
        return h == hint_host or h.endswith("." + hint_host)

    # URLs from grounding metadata (real indexed URLs)
    for ch in chunks:
        uri = (ch.get("web", {}) or {}).get("uri", "") if isinstance(ch, dict) else ""
        if not uri:
            continue
        try:
            real = session.get(uri, allow_redirects=True, timeout=timeout).url
            real = clean_url(real)
        except Exception:
            real = clean_url(uri)
        if real and real not in seen:
            host = urlparse(real).netloc.lower()
            if _t2_host_ok(host) and not any(bad in host for bad in BAD_HOSTS) and not is_pdf_or_file(real):
                seen.add(real)
                candidates.append(Candidate(url=real, source="grounding"))

    # URLs mentioned in text response
    for raw in re.findall(r"https?://[^\s\]\)\"'<>]+", text_response or ""):
        url = clean_url(raw.rstrip(".,;:"))
        if url and url not in seen:
            host = urlparse(url).netloc.lower()
            if _t2_host_ok(host) and not any(bad in host for bad in BAD_HOSTS) and not is_pdf_or_file(url):
                seen.add(url)
                candidates.append(Candidate(url=url, source="grounding"))

    print(f"      grounding: {len(chunks)} chunks, {len(candidates)} candidate URLs", flush=True)

    # Fetch each to get content for Tier 3
    for c in candidates:
        if is_broad_landing(c.url):
            c.fetchable = False
            continue
        page = fetch_page(session, c.url, timeout)
        if page.ok and not is_soft_404(page):
            c.page = page
            c.content_preview = page.text[:1200]
            c.fetchable = True
        else:
            c.fetchable = False

    return candidates


def tier2_find_site_grounded(session: requests.Session, model: str,
                             municipio: str, timeout: int = 15) -> Page | None:
    """Discover the official prefeitura domain via grounded search.

    Fallback for when the free slug guesses (Tier 0) miss because the real
    host is non-obvious (abbreviations like pmpf, shortened names like
    caxias, geo-blocked sites, migrations). No fixed rules: Gemini + Google
    find the domain, and we validate it with the same score bar as Tier 0.
    """
    prompt = (
        f"Qual e o site OFICIAL da Prefeitura Municipal de {municipio} "
        f"({UF_NOME}, {UF_SIGLA}, Brasil)?\n"
        "Responda com a URL da PAGINA INICIAL oficial (dominio .rs.gov.br, "
        ".atende.net ou outro dominio oficial da prefeitura). "
        "Nao responda com redes sociais, wikipedia, noticias nem portais de terceiros."
    )
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {"temperature": 0.0, "maxOutputTokens": 1024},
    }
    try:
        data = gemini_post(session, model, payload, timeout=90)
    except Exception as e:
        print(f"      grounded site discovery error: {e}", flush=True)
        return None

    cand = (data.get("candidates") or [{}])[0]
    text_response = "\n".join(
        p.get("text", "") for p in (cand.get("content", {}) or {}).get("parts", [])
        if isinstance(p, dict)
    )
    chunks = (cand.get("groundingMetadata", {}) or {}).get("groundingChunks", []) or []

    # Collect candidate homepage URLs: grounding chunks first (real indexed
    # URLs), then any URL mentioned in the text answer.
    raw_urls: list[str] = []
    for ch in chunks:
        uri = (ch.get("web", {}) or {}).get("uri", "") if isinstance(ch, dict) else ""
        if uri:
            try:
                raw_urls.append(session.get(uri, allow_redirects=True, timeout=timeout).url)
            except Exception:
                raw_urls.append(uri)
    raw_urls.extend(re.findall(r"https?://[^\s\]\)\"'<>]+", text_response or ""))

    # Reduce to candidate base domains, skipping junk hosts.
    seen: set[str] = set()
    base_urls: list[str] = []
    for raw in raw_urls:
        url = clean_url((raw or "").rstrip(".,;:"))
        host = urlparse(url).netloc.lower()
        if not host or host in seen:
            continue
        if any(bad in host for bad in BAD_HOSTS):
            continue
        seen.add(host)
        base_urls.append(f"{urlparse(url).scheme}://{host}/")

    print(f"      grounded site discovery: {len(base_urls)} domain candidates", flush=True)

    best = None
    best_score = -1
    for url in base_urls:
        page = fetch_page(session, url, timeout)
        if not page.ok:
            continue
        score = score_site_page(page, municipio)
        if score > best_score:
            best_score = score
            best = page
    if best and best_score >= 5:
        migrated = _check_migration(session, best, timeout)
        return migrated or best
    return None


def tier2_directed_bucket_search(session: requests.Session, model: str,
                                 municipio: str, host: str,
                                 bucket_name: str,
                                 timeout: int = 15) -> list[Candidate]:
    """Targeted grounding search for a specific missing bucket on a known host."""
    prompt = (
        f"Encontre a pagina INDICE/LISTAGEM de {bucket_name} da Prefeitura de "
        f"{municipio} ({UF_NOME}, {UF_SIGLA}, Brasil).\n"
        f"Busque: {bucket_name} site:{host}\n"
        "Queremos a pagina que LISTA VARIOS editais/processos, "
        "NAO um edital individual nem PDF.\n"
        "Liste as URLs encontradas."
    )
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {"temperature": 0.0, "maxOutputTokens": 1024},
    }
    try:
        data = gemini_post(session, model, payload, timeout=90)
    except Exception as e:
        print(f"      directed search error: {e}", flush=True)
        return []

    cand = (data.get("candidates") or [{}])[0]
    text_response = "\n".join(
        p.get("text", "") for p in (cand.get("content", {}) or {}).get("parts", [])
        if isinstance(p, dict)
    )
    chunks = (cand.get("groundingMetadata", {}) or {}).get("groundingChunks", []) or []

    candidates = []
    seen: set[str] = set()
    host_base = host.lower().lstrip("www.")

    def _host_ok(h: str) -> bool:
        h = h.lower().lstrip("www.")
        return h == host_base or h.endswith("." + host_base)

    for ch in chunks:
        uri = (ch.get("web", {}) or {}).get("uri", "") if isinstance(ch, dict) else ""
        if not uri:
            continue
        try:
            real = session.get(uri, allow_redirects=True, timeout=timeout).url
            real = clean_url(real)
        except Exception:
            real = clean_url(uri)
        if real and real not in seen:
            h = urlparse(real).netloc.lower()
            if _host_ok(h) and not any(bad in h for bad in BAD_HOSTS) and not is_pdf_or_file(real):
                seen.add(real)
                candidates.append(Candidate(url=real, source="directed_grounding"))

    for raw in re.findall(r"https?://[^\s\]\)\"'<>]+", text_response or ""):
        url = clean_url(raw.rstrip(".,;:"))
        if url and url not in seen:
            h = urlparse(url).netloc.lower()
            if _host_ok(h) and not any(bad in h for bad in BAD_HOSTS) and not is_pdf_or_file(url):
                seen.add(url)
                candidates.append(Candidate(url=url, source="directed_grounding"))

    print(f"      directed: {len(candidates)} candidates for {bucket_name}", flush=True)

    for c in candidates:
        if is_broad_landing(c.url):
            c.fetchable = False
            continue
        page = fetch_page(session, c.url, timeout)
        if page.ok and not is_soft_404(page):
            c.page = page
            c.content_preview = page.text[:1200]
            c.fetchable = True
        else:
            c.fetchable = False

    return candidates


# ---------------------------------------------------------------------------
# TIER 3: Gemini verifier/selector (discrete decisions, no scores)
# ---------------------------------------------------------------------------
TIER3_DECISIONS = [
    "indice_oficial",
    "indice_oficial_combinado",
    "portal_externo_oficial",
    "detalle_individual_rechazado",
    "licitacao_rechazada",
    "concurso_cultural_rechazado",
    "pagina_generica_rechazada",
    "nao_encontrado",
    "revisar",
]


def tier3_classify_and_pick(session: requests.Session, model: str,
                            municipio: str, candidates: list[Candidate],
                            timeout: int = 30) -> dict[str, str]:
    """Send all candidates to Gemini for classification and selection.

    Returns dict with keys 'url_concursos' and 'url_processos_seletivos',
    each either a URL string or empty string.
    """
    if not candidates:
        return {"url_concursos": "", "url_processos_seletivos": "", "razao": ""}

    fetchable = [c for c in candidates if c.fetchable and c.page]
    if not fetchable:
        return {"url_concursos": "", "url_processos_seletivos": "", "razao": ""}

    items = []
    for i, c in enumerate(fetchable[:15]):
        preview = re.sub(r"[\x00-\x1f]+", " ", c.content_preview[:600])
        items.append({
            "id": i,
            "url": c.url,
            "menu_text": c.menu_text or "(encontrado via busca)",
            "source": c.source,
            "title": (c.page.title if c.page else "")[:120],
            "content_preview": preview,
        })

    prompt = (
        f"Prefeitura de {municipio} ({UF_NOME}, {UF_SIGLA}). "
        f"Analise {len(items)} URLs candidatas e escolha a melhor pagina-INDICE "
        f"(listagem de VARIOS editais) para concursos publicos e para processos seletivos (PSS).\n\n"
        "Regras:\n"
        "- Queremos pagina INDICE/LISTAGEM, NAO edital individual, PDF ou noticia.\n"
        "- Prefira o INDICE CANONICO: a listagem mais ampla e estavel. Entre uma\n"
        "  vista de TODOS os anos e uma filtrada por um ano so, escolha a de todos\n"
        "  os anos.\n"
        "- Se existem paginas SEPARADAS para concursos e para PSS, use a especifica\n"
        "  de cada bucket; so use uma pagina combinada se nao houver separadas.\n"
        "- Uma pagina de CATEGORIA especifica (ex: /concurso/categoria/25/concurso)\n"
        "  e MELHOR que a pagina raiz generica (/concurso) porque filtra exatamente\n"
        "  o tipo desejado, mesmo que tenha menos itens.\n"
        "- Se duas sao parecidas, escolha a mais completa e atualizada.\n"
        "- Rejeite licitacao/pregao/compras e concurso cultural (soberanas/rainhas).\n"
        "- Se nenhuma serve, deixe vazio. NAO invente URLs.\n\n"
        "Candidatos:\n"
    )
    for item in items:
        prompt += (
            f"  [{item['id']}] {item['url']}\n"
            f"      menu: {item['menu_text']}\n"
            f"      title: {item['title']}\n"
            f"      preview: {item['content_preview'][:300]}\n\n"
        )
    prompt += (
        "Responda JSON com as URLs PRIMEIRO e a razao por ultimo "
        "(uma frase curta, max 20 palavras): "
        "{\"url_concursos\": \"url ou vazio\", "
        "\"url_processos_seletivos\": \"url ou vazio\", \"razao\": \"curto\"}"
    )

    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.0, "maxOutputTokens": 4096,
            "responseMimeType": "application/json",
        },
    }
    data = gemini_post(session, model, payload, timeout=60)
    try:
        text = "\n".join(
            p.get("text", "") for p in data["candidates"][0]["content"]["parts"]
        )
        # Try direct parse; if truncated, attempt repair
        try:
            result = json.loads(text)
        except json.JSONDecodeError:
            # Try extracting URLs even from truncated JSON
            url_c_m = re.search(r'"url_concursos"\s*:\s*"(https?://[^"]+)"', text)
            url_p_m = re.search(r'"url_processos_seletivos"\s*:\s*"(https?://[^"]+)"', text)
            if url_c_m or url_p_m:
                result = {
                    "url_concursos": url_c_m.group(1) if url_c_m else "",
                    "url_processos_seletivos": url_p_m.group(1) if url_p_m else "",
                }
                print(f"      tier3: recovered from truncated JSON", flush=True)
            else:
                print(f"      tier3: no valid JSON: {text[:300]}", flush=True)
                return {"url_concursos": "", "url_processos_seletivos": "", "razao": ""}

        url_c = result.get("url_concursos", "")
        url_p = result.get("url_processos_seletivos", "")

        # Validate that chosen URLs are actually in our candidate list
        valid_urls = {c.url for c in fetchable}
        if url_c and url_c not in valid_urls:
            print(f"      ! url_concursos not in candidates, rejected: {url_c[:80]}", flush=True)
            url_c = ""
        if url_p and url_p not in valid_urls:
            print(f"      ! url_processos not in candidates, rejected: {url_p[:80]}", flush=True)
            url_p = ""

        razao = result.get("razao", "")
        if url_c:
            print(f"      → concursos: {url_c}", flush=True)
        if url_p:
            print(f"      → processos: {url_p}", flush=True)
        if razao:
            print(f"      razao: {razao}", flush=True)
        if not url_c and not url_p:
            print(f"      → nenhuma URL valida", flush=True)

        return {"url_concursos": url_c, "url_processos_seletivos": url_p,
                "razao": razao}

    except Exception as e:
        print(f"      tier3 parse error: {e}", flush=True)
        return {"url_concursos": "", "url_processos_seletivos": "", "razao": ""}


# ---------------------------------------------------------------------------
# TIER 4: Playwright directed navigation
# ---------------------------------------------------------------------------
_BROWSER = None


def _find_chromium() -> str | None:
    import glob
    for pat in ("/opt/pw-browsers/chromium-*/chrome-linux/chrome",
                "/opt/pw-browsers/chromium-*/chrome-linux/headless_shell"):
        hits = sorted(glob.glob(pat))
        if hits:
            return hits[-1]
    return None


def _get_browser():
    global _BROWSER
    if _BROWSER is None:
        from playwright.sync_api import sync_playwright
        pw = sync_playwright().start()
        chrome_path = _find_chromium()
        # Hardening for proxied/headless environments. Secure DNS (DoH) probes
        # to dns.google are unreachable behind an egress proxy and flood the
        # net stack, so turn DoH and background networking off and let the
        # proxy resolve names via CONNECT.
        launch_args = [
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--dns-over-https-mode=off",
            "--disable-features=DnsOverHttps,DnsOverHttpsUpgrade",
            "--disable-background-networking",
            "--disable-component-update",
            "--no-pings",
        ]
        proxy_url = os.environ.get("HTTPS_PROXY") or os.environ.get("https_proxy")
        if proxy_url:
            launch_args.append(f"--proxy-server={proxy_url}")
            # A re-terminating egress proxy forges a per-host cert chain and
            # cannot complete BoringSSL's TLS 1.3 ClientHello (post-quantum
            # keyshare). Accept the forged cert and cap at TLS 1.2 so the
            # tunnel handshake succeeds. Only applied when a proxy is present;
            # direct-egress (production) keeps modern TLS.
            launch_args += [
                "--ignore-certificate-errors",
                "--test-type",
                "--ssl-version-max=tls1.2",
            ]
        _BROWSER = pw.chromium.launch(
            headless=True,
            executable_path=chrome_path,
            args=launch_args,
        )
    return _BROWSER


def tier4_playwright_collect(url: str, municipio: str) -> list[Candidate]:
    """Navigate the site like a human: open menus, follow relevant links."""
    try:
        browser = _get_browser()
    except Exception as e:
        print(f"      playwright unavailable: {e}", flush=True)
        return []

    candidates = []
    all_keywords = []
    for kws in BUCKET_KEYWORDS.values():
        all_keywords.extend(kws)

    try:
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
            ignore_https_errors=True,
        )
        page = context.new_page()
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(2000)

        # Extract all links from the rendered page
        links = page.evaluate("""() => {
            const results = [];
            document.querySelectorAll('a[href]').forEach(el => {
                results.push({href: el.href, text: (el.innerText || '').trim()});
            });
            return results;
        }""")

        seen = set()
        relevant_links = []
        container_links = []

        for link in links:
            href = link.get("href", "")
            text = link.get("text", "")
            if not href.startswith("http") or href in seen:
                continue
            host = urlparse(href).netloc.lower()
            if any(bad in host for bad in BAD_HOSTS):
                continue
            text_n = norm(text)
            href_n = norm(unquote(urlparse(href).path))

            if any(kw in text_n or kw in href_n for kw in all_keywords):
                seen.add(href)
                relevant_links.append((href, text))
            elif any(kw in text_n or kw in href_n for kw in CONTAINER_KEYWORDS):
                container_links.append((href, text))

        # Follow container links one level deep
        for container_href, container_text in container_links[:5]:
            if container_href in seen:
                continue
            try:
                page2 = context.new_page()
                page2.goto(container_href, wait_until="domcontentloaded", timeout=15000)
                page2.wait_for_timeout(1500)
                sub_links = page2.evaluate("""() => {
                    const results = [];
                    document.querySelectorAll('a[href]').forEach(el => {
                        results.push({href: el.href, text: (el.innerText || '').trim()});
                    });
                    return results;
                }""")
                for sl in sub_links:
                    sh = sl.get("href", "")
                    st = sl.get("text", "")
                    if not sh.startswith("http") or sh in seen:
                        continue
                    st_n = norm(st)
                    sh_n = norm(unquote(urlparse(sh).path))
                    if any(kw in st_n or kw in sh_n for kw in all_keywords):
                        seen.add(sh)
                        relevant_links.append((sh, f"{container_text} > {st}"))
                page2.close()
            except Exception:
                pass

        for href, text in relevant_links:
            if is_broad_landing(href) or is_pdf_or_file(href):
                continue
            candidates.append(Candidate(
                url=href, source="playwright", menu_text=text,
            ))

        page.close()
        context.close()
    except Exception as e:
        print(f"      playwright error: {e}", flush=True)

    print(f"      playwright: {len(candidates)} candidates found", flush=True)

    # Fetch content for Tier 3 classification
    session = make_session()
    for c in candidates:
        pg = fetch_page(session, c.url, timeout=12)
        if pg.ok and not is_soft_404(pg):
            c.page = pg
            c.content_preview = pg.text[:1200]
            c.fetchable = True
        else:
            c.fetchable = False

    return candidates


# ---------------------------------------------------------------------------
# Combined-page detection helper
# ---------------------------------------------------------------------------
CONCURSO_SIGNALS = ["concurso publico", "concursos publicos", "concurso público"]
PSS_SIGNALS = [
    "processo seletivo", "processos seletivos", "seletivo simplificado",
    "selecao publica", "seleção pública", "pss",
]


def _try_combined_fill(session: requests.Session, chosen: dict,
                       bucket_tier: dict, razones: list,
                       candidates: list[Candidate]) -> None:
    """If one bucket has a URL and the other doesn't, check if the filled
    bucket's page content mentions both types — making it a combined page."""
    filled_key = empty_key = None
    if chosen["url_concursos"] and not chosen["url_processos_seletivos"]:
        filled_key, empty_key = "url_concursos", "url_processos_seletivos"
        signals = PSS_SIGNALS
    elif chosen["url_processos_seletivos"] and not chosen["url_concursos"]:
        filled_key, empty_key = "url_processos_seletivos", "url_concursos"
        signals = CONCURSO_SIGNALS
    else:
        return

    filled_url = chosen[filled_key]
    page = None
    for c in candidates:
        if c.url == filled_url and c.page:
            page = c.page
            break
    if not page:
        return

    content = norm(page.title + " " + page.text[:3000])
    if any(s in content for s in signals):
        chosen[empty_key] = filled_url
        bucket_tier[empty_key] = bucket_tier.get(filled_key, "") + "_combined"
        razones.append(f"[combined] Page also contains {empty_key.split('_')[1]} content")
        print(f"      combined: {empty_key} filled from {filled_key}", flush=True)


# ---------------------------------------------------------------------------
# Main pipeline: process one municipality
# ---------------------------------------------------------------------------
def _assign_confidence(tier: str) -> str:
    """Initial confidence from tier. May be upgraded later by verification."""
    if not tier:
        return ""
    if tier in ("t1",):
        return "confirmado"
    if tier in ("t2", "t2dir", "t4"):
        return "probable"
    if tier.endswith("_combined"):
        return "revisar"
    return "probable"


LISTING_SIGNALS = [
    r"\b\d{1,3}/20[12]\d\b",       # edital numbers like 001/2024
    r"edital\s+n",                  # "Edital Nº"
    r"inscri[cç][oõ]es\s+(aberta|encerrada)",
    r"resultado\s+(final|parcial|preliminar)",
    r"homologa[cç][aã]o",
    r"retifica[cç][aã]o",
]
LISTING_RE = re.compile("|".join(LISTING_SIGNALS), re.I)

CONCURSO_VERIFY_KW = [
    "concurso publico", "concursos publicos", "concurso público",
    "concursos públicos",
]
PSS_VERIFY_KW = [
    "processo seletivo", "processos seletivos", "seletivo simplificado",
    "selecao publica", "seleção pública", "pss ",
]


def _deterministic_verify(url: str, bucket: str,
                          all_candidates: list[Candidate]) -> bool:
    """Check if URL content looks like a real listing page for this bucket.

    Returns True if confident enough to upgrade probable→confirmado.
    """
    cand = next((c for c in all_candidates if c.url == url and c.page), None)
    if not cand or not cand.page:
        return False
    page = cand.page
    text = norm(page.title + " " + page.text[:2000])

    kw_list = CONCURSO_VERIFY_KW if bucket == "concursos" else PSS_VERIFY_KW
    has_keyword = any(k in text for k in kw_list)
    listing_matches = len(LISTING_RE.findall(page.text[:3000]))
    has_multiple_items = listing_matches >= 2

    return has_keyword and has_multiple_items


def batch_gemini_verify(session: requests.Session, model: str,
                        to_verify: list[dict],
                        timeout: int = 30) -> dict[str, str]:
    """Verify uncertain URLs in a single Gemini call.

    to_verify: list of {"municipio": str, "bucket": str, "url": str,
                        "title": str, "preview": str}
    Returns {f"{municipio}|{bucket}": "confirmado" or "revisar"}
    """
    if not to_verify or not gemini_api_key():
        return {}

    items_text = ""
    for i, item in enumerate(to_verify[:30]):
        items_text += (
            f"[{i}] Municipio: {item['municipio']}, Bucket: {item['bucket']}\n"
            f"    Site oficial: {item.get('site_base', '')}\n"
            f"    URL: {item['url']}\n"
            f"    Titulo: {item['title'][:120]}\n"
            f"    Preview: {item['preview'][:250]}\n\n"
        )

    prompt = (
        "Voce e um verificador de paginas de concursos publicos municipais.\n"
        "Para cada item abaixo, responda se a URL e uma pagina INDICE/LISTAGEM "
        "valida para o bucket indicado (concursos ou processos seletivos).\n\n"
        "Criterios para CONFIRMAR:\n"
        "- A pagina lista MULTIPLOS editais/concursos/processos (nao so um)\n"
        "- O conteudo corresponde ao bucket (concursos OU processos seletivos)\n"
        "- E uma pagina de listagem, nao um edital individual ou PDF\n"
        "- Paginas combinadas (ambos tipos) sao validas para ambos buckets\n"
        "- A URL pertence ao MESMO dominio do site oficial (ou subdominio)\n\n"
        "Criterios para REVISAR:\n"
        "- Pagina de um unico edital\n"
        "- Conteudo nao corresponde ao bucket\n"
        "- Pagina generica sem editais visiveis\n"
        "- Licitacoes ou concursos culturais\n"
        "- URL de dominio DIFERENTE do site oficial (ex: banca, fundacao, outro orgao)\n\n"
        f"Items a verificar:\n{items_text}\n"
        "Responda JSON array. Cada elemento: "
        "{\"id\": N, \"veredicto\": \"confirmado\" ou \"revisar\", "
        "\"motivo\": \"frase curta\"}\n"
    )

    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.0, "maxOutputTokens": 4096,
            "responseMimeType": "application/json",
        },
    }
    try:
        data = gemini_post(session, model, payload, timeout=60)
        text = "\n".join(
            p.get("text", "") for p in data["candidates"][0]["content"]["parts"]
        )
        verdicts = json.loads(text)
        if isinstance(verdicts, dict) and "items" in verdicts:
            verdicts = verdicts["items"]
        if not isinstance(verdicts, list):
            verdicts = [verdicts]

        result = {}
        for v in verdicts:
            idx = v.get("id", -1)
            if 0 <= idx < len(to_verify):
                item = to_verify[idx]
                key = f"{item['municipio']}|{item['bucket']}"
                raw_veredicto = v.get("veredicto", "revisar").lower().strip()
                veredicto = "confirmado" if raw_veredicto.startswith("confirm") else "revisar"
                motivo = v.get("motivo", "")
                result[key] = (veredicto, motivo)
                print(f"    verify [{idx}] {item['municipio']}/{item['bucket']}: "
                      f"{veredicto} — {motivo}", flush=True)
        return result
    except Exception as e:
        print(f"    batch verify error: {e}", flush=True)
        return {}


@dataclass
class MunicipioResult:
    municipio: str
    site_base: str = ""
    url_concursos: str = ""
    url_processos_seletivos: str = ""
    method: str = ""
    notes: str = ""
    tier_concursos: str = ""
    tier_processos: str = ""
    razao: str = ""
    confianza_concursos: str = ""
    confianza_processos: str = ""
    urls_extras_concursos: str = ""
    urls_extras_processos: str = ""


def process_municipio(session: requests.Session, municipio: str,
                      model: str, timeout: int = 15,
                      use_playwright: bool = True) -> MunicipioResult:
    result = MunicipioResult(municipio=municipio)
    tiers_used = []
    all_candidates: list[Candidate] = []

    # --- TIER 0: Find site base (free slug guesses) ---
    print(f"  [{municipio}] Tier 0: finding site...", flush=True)
    home = tier0_find_site(session, municipio, timeout)
    if home:
        tiers_used.append("t0")
    elif gemini_api_key():
        # Free path missed (non-obvious domain, geo-block, migration):
        # let grounded search discover the official domain.
        print(f"    Tier 0 free miss; grounded site discovery...", flush=True)
        home = tier2_find_site_grounded(session, model, municipio, timeout)
        if home:
            tiers_used.append("t2site")

    if not home:
        result.notes = "site_not_found"
        result.method = "+".join(tiers_used) + ("+" if tiers_used else "") + "tier0_failed"
        return result
    result.site_base = clean_url(home.url)
    print(f"    site: {result.site_base}", flush=True)

    # --- TIER 1: Free link discovery ---
    print(f"    Tier 1: scanning links...", flush=True)
    t1_candidates = tier1_collect_candidates(session, home, municipio, timeout)
    all_candidates.extend(t1_candidates)
    fetchable_t1 = [c for c in t1_candidates if c.fetchable]
    print(f"    Tier 1: {len(fetchable_t1)} fetchable candidates from {len(t1_candidates)} found", flush=True)
    tiers_used.append("t1")

    # --- TIER 3 on Tier 1 candidates (if we have any) ---
    chosen = {"url_concursos": "", "url_processos_seletivos": ""}
    bucket_tier = {"url_concursos": "", "url_processos_seletivos": ""}
    razones: list[str] = []

    def _record(picked: dict, tier_label: str) -> None:
        """Fill empty buckets from a Tier 3 result and note which tier won."""
        for key in ("url_concursos", "url_processos_seletivos"):
            if not chosen[key] and picked.get(key):
                chosen[key] = picked[key]
                bucket_tier[key] = tier_label
        if picked.get("razao"):
            razones.append(f"[{tier_label}] {picked['razao']}")

    if fetchable_t1 and gemini_api_key():
        print(f"    Tier 3: classifying {len(fetchable_t1)} Tier 1 candidates...", flush=True)
        picked = tier3_classify_and_pick(session, model, municipio, fetchable_t1, timeout)
        _record(picked, "t1")
        tiers_used.append("t3")

    # --- Combined-page detection ---
    # If Tier 3 filled one bucket but not the other, check if the chosen
    # page's content mentions both types. If so, it's a combined page.
    _try_combined_fill(session, chosen, bucket_tier, razones, all_candidates)

    # --- TIER 2: Grounded search (only for missing buckets) ---
    missing_buckets = []
    if not chosen["url_concursos"]:
        missing_buckets.append("concursos publicos")
    if not chosen["url_processos_seletivos"]:
        missing_buckets.append("processos seletivos")

    if missing_buckets and gemini_api_key():
        print(f"    Tier 2: grounded search (missing: {', '.join(missing_buckets)})...", flush=True)
        try:
            t2_candidates = tier2_grounded_search(
                session, model, municipio, result.site_base, timeout,
            )
            # Filter out candidates we already have
            existing_urls = {c.url for c in all_candidates}
            new_t2 = [c for c in t2_candidates if c.url not in existing_urls]
            all_candidates.extend(new_t2)
            tiers_used.append("t2")

            # Run Tier 3 on new candidates (plus any unfilled from before)
            fetchable_new = [c for c in new_t2 if c.fetchable]
            if fetchable_new:
                print(f"    Tier 3: classifying {len(fetchable_new)} grounded candidates...", flush=True)
                picked = tier3_classify_and_pick(session, model, municipio, fetchable_new, timeout)
                _record(picked, "t2")
        except Exception as e:
            print(f"    Tier 2 error: {e}", flush=True)
            tiers_used.append("t2_err")

    # Combined-page check again after Tier 2
    _try_combined_fill(session, chosen, bucket_tier, razones, all_candidates)

    # --- Directed grounding per bucket ---
    # If general grounding didn't find a bucket, search specifically for it
    # on the known host (e.g. "processos seletivos site:pmpf.rs.gov.br").
    dir_missing = []
    if not chosen["url_concursos"]:
        dir_missing.append(("url_concursos", "concursos publicos"))
    if not chosen["url_processos_seletivos"]:
        dir_missing.append(("url_processos_seletivos", "processos seletivos"))

    if dir_missing and result.site_base and gemini_api_key():
        host = urlparse(result.site_base).netloc
        for bucket_key, bucket_name in dir_missing:
            print(f"    Directed grounding: {bucket_name} on {host}...", flush=True)
            try:
                t2d = tier2_directed_bucket_search(
                    session, model, municipio, host, bucket_name, timeout,
                )
                existing_urls = {c.url for c in all_candidates}
                new_d = [c for c in t2d if c.url not in existing_urls]
                all_candidates.extend(new_d)
                fetchable_d = [c for c in new_d if c.fetchable]
                if fetchable_d:
                    picked = tier3_classify_and_pick(
                        session, model, municipio, fetchable_d, timeout,
                    )
                    _record(picked, "t2dir")
                tiers_used.append("t2dir")
            except Exception as e:
                print(f"    Directed grounding error: {e}", flush=True)

    # Combined check after directed grounding
    _try_combined_fill(session, chosen, bucket_tier, razones, all_candidates)

    # --- TIER 4: Playwright (last resort for still-missing buckets) ---
    still_missing = []
    if not chosen["url_concursos"]:
        still_missing.append("concursos")
    if not chosen["url_processos_seletivos"]:
        still_missing.append("processos")

    if still_missing and use_playwright:
        print(f"    Tier 4: playwright navigation (missing: {', '.join(still_missing)})...", flush=True)
        try:
            t4_candidates = tier4_playwright_collect(result.site_base, municipio)
            existing_urls = {c.url for c in all_candidates}
            new_t4 = [c for c in t4_candidates if c.url not in existing_urls]
            all_candidates.extend(new_t4)
            tiers_used.append("t4")

            fetchable_t4 = [c for c in new_t4 if c.fetchable]
            if fetchable_t4 and gemini_api_key():
                print(f"    Tier 3: classifying {len(fetchable_t4)} playwright candidates...", flush=True)
                picked = tier3_classify_and_pick(session, model, municipio, fetchable_t4, timeout)
                _record(picked, "t4")
        except Exception as e:
            print(f"    Tier 4 error: {e}", flush=True)

    # Final combined-page check after all tiers
    _try_combined_fill(session, chosen, bucket_tier, razones, all_candidates)

    # --- Assemble result ---
    result.url_concursos = chosen.get("url_concursos", "")
    result.url_processos_seletivos = chosen.get("url_processos_seletivos", "")
    result.method = "+".join(tiers_used)
    result.tier_concursos = bucket_tier["url_concursos"]
    result.tier_processos = bucket_tier["url_processos_seletivos"]
    result.razao = " | ".join(razones)

    # --- Confidence assignment ---
    result.confianza_concursos = _assign_confidence(result.tier_concursos)
    result.confianza_processos = _assign_confidence(result.tier_processos)

    # Deterministic upgrade: probable→confirmado if content clearly matches
    if result.confianza_concursos == "probable" and result.url_concursos:
        if _deterministic_verify(result.url_concursos, "concursos", all_candidates):
            result.confianza_concursos = "confirmado"
    if result.confianza_processos == "probable" and result.url_processos_seletivos:
        if _deterministic_verify(result.url_processos_seletivos, "processos", all_candidates):
            result.confianza_processos = "confirmado"

    # Downgrade to "revisar" when site not found or all tiers exhausted
    if not result.url_concursos and not result.url_processos_seletivos:
        if any(c.fetchable for c in all_candidates):
            result.confianza_concursos = "revisar"
            result.confianza_processos = "revisar"

    # --- Collect extra valid URLs (others Tier 3 could have picked) ---
    # Any fetchable candidate on the same host that wasn't chosen
    chosen_urls = {result.url_concursos, result.url_processos_seletivos}
    concurso_kw = {"concurso", "concursos"}
    pss_kw = {"seletivo", "seletivos", "pss", "selecao"}
    extras_c: list[str] = []
    extras_p: list[str] = []
    for c in all_candidates:
        if not c.fetchable or c.url in chosen_urls:
            continue
        text_lower = (c.menu_text or "").lower() + " " + c.url.lower()
        if c.page and c.page.title:
            text_lower += " " + c.page.title.lower()
        if any(k in text_lower for k in concurso_kw):
            extras_c.append(c.url)
        if any(k in text_lower for k in pss_kw):
            extras_p.append(c.url)
    result.urls_extras_concursos = " | ".join(extras_c[:5])
    result.urls_extras_processos = " | ".join(extras_p[:5])

    notes_parts = []
    total = len(all_candidates)
    fetchable = len([c for c in all_candidates if c.fetchable])
    if total > 0:
        notes_parts.append(f"{total} candidates ({fetchable} fetchable)")
    if not result.url_concursos and not result.url_processos_seletivos:
        notes_parts.append("no valid index page found")
    result.notes = "; ".join(notes_parts)

    return result


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
OUTPUT_FIELDS = [
    "uf", "municipio", "site_base",
    "url_concursos", "confianza_concursos",
    "url_processos_seletivos", "confianza_processos",
    "urls_extras_concursos", "urls_extras_processos",
    "tier_concursos", "tier_processos",
    "method", "razao", "notes", "checked_at",
]


def write_results(results: list[MunicipioResult], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    now = datetime.now(timezone.utc).isoformat()
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        for r in results:
            writer.writerow({
                "uf": UF_SIGLA,
                "municipio": r.municipio,
                "site_base": r.site_base,
                "url_concursos": r.url_concursos,
                "confianza_concursos": r.confianza_concursos,
                "url_processos_seletivos": r.url_processos_seletivos,
                "confianza_processos": r.confianza_processos,
                "urls_extras_concursos": r.urls_extras_concursos,
                "urls_extras_processos": r.urls_extras_processos,
                "tier_concursos": r.tier_concursos,
                "tier_processos": r.tier_processos,
                "method": r.method,
                "razao": r.razao,
                "notes": r.notes,
                "checked_at": now,
            })
    print(f"\nCSV written to {path}", flush=True)

    # --- Excel output with proper formatting ---
    xlsx_path = path.with_suffix(".xlsx")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Concursos RS"

        # Colors for confidence levels
        fills = {
            "confirmado": PatternFill("solid", fgColor="C6EFCE"),  # green
            "probable":   PatternFill("solid", fgColor="FFEB9C"),  # yellow
            "revisar":    PatternFill("solid", fgColor="FFC7CE"),  # red/pink
        }
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        link_font = Font(color="0563C1", underline="single", size=10)
        wrap_align = Alignment(wrap_text=True, vertical="top")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # User-friendly column names
        excel_cols = [
            ("Municipio", 22),
            ("Site Base", 30),
            ("URL Concursos", 45),
            ("Confianza C", 14),
            ("URL Processos Seletivos", 45),
            ("Confianza P", 14),
            ("URLs Extras Concursos", 40),
            ("URLs Extras Processos", 40),
            ("Tier C", 8),
            ("Tier P", 8),
            ("Razon IA", 60),
            ("Notas", 35),
            ("Fecha", 22),
        ]
        for col_idx, (name, width) in enumerate(excel_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = width

        for row_idx, r in enumerate(results, 2):
            vals = [
                r.municipio, r.site_base,
                r.url_concursos, r.confianza_concursos,
                r.url_processos_seletivos, r.confianza_processos,
                r.urls_extras_concursos, r.urls_extras_processos,
                r.tier_concursos, r.tier_processos,
                r.razao, r.notes, now,
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = wrap_align
                cell.border = thin_border

            # Color-code confidence columns
            for conf_col in (4, 6):
                cell = ws.cell(row=row_idx, column=conf_col)
                if cell.value in fills:
                    cell.fill = fills[cell.value]

            # Make URLs clickable
            for url_col in (2, 3, 5):
                cell = ws.cell(row=row_idx, column=url_col)
                if cell.value and str(cell.value).startswith("http"):
                    cell.font = link_font
                    cell.hyperlink = str(cell.value)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
        print(f"Excel written to {xlsx_path}", flush=True)
    except ImportError:
        print("openpyxl not installed, skipping Excel output", flush=True)
    except Exception as e:
        print(f"Excel write error: {e}", flush=True)


# ---------------------------------------------------------------------------
# Municipality loading
# ---------------------------------------------------------------------------
def load_municipios_from_golden(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [row["municipio"] for row in reader if row.get("municipio")]


def load_municipios_from_tce(session: requests.Session,
                             timeout: int = 30) -> list[str]:
    try:
        resp = session.get(DEFAULT_MUNICIPIOS_URL, timeout=timeout)
        resp.raise_for_status()
        reader = csv.DictReader(resp.text.splitlines(), delimiter=";")
        return sorted(set(
            row["NOME_MUNICIPIO"]
            for row in reader
            if (row.get("UF") or row.get("SIGLA_UF", "")) == "RS"
            and row.get("NOME_MUNICIPIO")
        ))
    except Exception as e:
        print(f"Error loading municipios: {e}", file=sys.stderr)
        return []


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="5-tier cascade for RS municipality resource discovery"
    )
    parser.add_argument("--golden", type=Path,
                        default=PROJECT_ROOT / "authority_first" / "data" / "golden_set_v1.csv",
                        help="Run only on golden set municipalities")
    parser.add_argument("--all", action="store_true",
                        help="Run on all 497 RS municipalities")
    parser.add_argument("--municipio", type=str,
                        help="Run on a single municipality")
    parser.add_argument("--output", type=Path,
                        default=PROJECT_ROOT / "data" / "cascade_output.csv")
    parser.add_argument("--model", type=str,
                        default=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"))
    parser.add_argument("--no-playwright", action="store_true",
                        help="Skip Tier 4")
    parser.add_argument("--timeout", type=int, default=15)
    parser.add_argument("--limit", type=int, default=0,
                        help="Limit number of municipalities")
    args = parser.parse_args()

    session = make_session()

    if args.municipio:
        municipios = [args.municipio]
    elif args.all:
        municipios = load_municipios_from_tce(session)
    else:
        municipios = load_municipios_from_golden(args.golden)

    if args.limit > 0:
        municipios = municipios[:args.limit]

    print(f"Processing {len(municipios)} municipalities", flush=True)
    print(f"Model: {args.model}", flush=True)
    print(f"Playwright: {'disabled' if args.no_playwright else 'enabled'}", flush=True)
    print("=" * 60, flush=True)

    results = []
    for i, muni in enumerate(municipios, 1):
        print(f"\n[{i}/{len(municipios)}] {muni}", flush=True)
        try:
            r = process_municipio(
                session, muni, args.model,
                timeout=args.timeout,
                use_playwright=not args.no_playwright,
            )
            results.append(r)
            status = []
            if r.url_concursos:
                status.append(f"C: {r.url_concursos}")
            if r.url_processos_seletivos:
                status.append(f"P: {r.url_processos_seletivos}")
            if not status:
                status.append("nothing found")
            print(f"  result → {'; '.join(status)}", flush=True)
        except Exception as e:
            print(f"  ERROR: {e}", flush=True)
            traceback.print_exc()
            results.append(MunicipioResult(municipio=muni, notes=f"error: {e}"))

    # --- Batch Gemini verification for uncertain results ---
    to_verify: list[dict] = []
    verify_index: dict[str, MunicipioResult] = {}
    for r in results:
        for bucket, url, conf in [
            ("concursos", r.url_concursos, r.confianza_concursos),
            ("processos", r.url_processos_seletivos, r.confianza_processos),
        ]:
            if conf == "probable" and url:
                to_verify.append({
                    "municipio": r.municipio, "bucket": bucket,
                    "url": url, "title": "", "preview": "",
                    "site_base": r.site_base,
                })
                verify_index[f"{r.municipio}|{bucket}"] = r

    if to_verify and gemini_api_key():
        print(f"\n{'='*60}", flush=True)
        print(f"Batch verification: {len(to_verify)} uncertain URLs", flush=True)
        # Re-fetch minimal content for verification
        for item in to_verify:
            try:
                pg = fetch_page(session, item["url"], timeout=args.timeout)
                item["title"] = pg.title[:150] if pg else ""
                item["preview"] = pg.text[:400] if pg else ""
            except Exception:
                pass

        verdicts = batch_gemini_verify(session, args.model, to_verify)
        for key, (veredicto, motivo) in verdicts.items():
            r = verify_index.get(key)
            if not r:
                continue
            muni, bucket = key.split("|", 1)
            if bucket == "concursos":
                r.confianza_concursos = veredicto
                if motivo:
                    r.notes += f"; verify_c: {motivo}"
            else:
                r.confianza_processos = veredicto
                if motivo:
                    r.notes += f"; verify_p: {motivo}"

        confirmed = sum(1 for _, (v, _) in verdicts.items() if v == "confirmado")
        print(f"  Verified: {confirmed}/{len(verdicts)} upgraded to confirmado",
              flush=True)

    write_results(results, args.output)

    # --- Summary ---
    found_c = sum(1 for r in results if r.url_concursos)
    found_p = sum(1 for r in results if r.url_processos_seletivos)
    conf_c = sum(1 for r in results if r.confianza_concursos == "confirmado")
    conf_p = sum(1 for r in results if r.confianza_processos == "confirmado")
    prob_c = sum(1 for r in results if r.confianza_concursos == "probable")
    prob_p = sum(1 for r in results if r.confianza_processos == "probable")
    rev_c = sum(1 for r in results if r.confianza_concursos == "revisar")
    rev_p = sum(1 for r in results if r.confianza_processos == "revisar")
    print(f"\nSummary: {found_c}/{len(results)} concursos, "
          f"{found_p}/{len(results)} processos found", flush=True)
    print(f"  Concursos  — confirmado: {conf_c}, probable: {prob_c}, revisar: {rev_c}",
          flush=True)
    print(f"  Processos  — confirmado: {conf_p}, probable: {prob_p}, revisar: {rev_p}",
          flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
