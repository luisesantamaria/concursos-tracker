#!/usr/bin/env python3
"""Cascade pipeline for municipality resource discovery (state-agnostic).

Finds the stable index/listing page for concursos and processos seletivos
in each municipality. Does NOT extract individual editals.

The TIERS are numbered by COST/TYPE, not by run order: Tier 2 (grounding) is
costlier than Tier 3 (verifier), so it actually runs AFTER Tier 3. What each does:
    Tier 0 — Site oficial: find/confirm the prefeitura's base domain.
    Tier 1 — Free link discovery: HTML menus, anchors, sitemap, transparência.
    Tier 2 — Grounded search: Gemini + Google Search (only if still missing).
    Tier 3 — Gemini selector: receives immutable, deterministically adjudicated
             CandidateRecords and returns a candidate_id for each bucket.
    Tier 4 — Playwright navigation agent: directed menu navigation as last resort.

ACTUAL RUN ORDER per municipality (each step runs ONLY if buckets remain empty —
spend expensive tools only when cheap ones fail):
    1. Tier 0           · Site oficial (free slug; grounded discovery if it misses)
    2. Tier 1           · Free links (renders the menu first if the site is a SPA)
    3. Tier 3           · AI picks among eligible Tier 1 CandidateRecords
    4. Tier 2           · Grounded Google search for the missing bucket -> Tier 3
    5. Directed grounding · "site:host {tipo}" per missing bucket -> Tier 3
    6. Tier 4           · Playwright navigates the menus as a human -> Tier 3
    7. FinalDecision derives directly from each exact SelectedResource. The
       compatibility batch adapts URL-only legacy rows through the same central
       adjudicator; selected records are never fetched or adjudicated again.

The 'method' field records which tiers fired (e.g. 't0+t1+t3'). UF/scope is set
by UF_SIGLA / UF_NOME below; the discovery logic itself is state-agnostic.
No numeric scorers. No magic constants. Discrete decisions + AI judgment.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import os
import random
import re
import sys
import time
import traceback
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import parse_qsl, urljoin, urlparse, unquote

import requests

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "scripts" / "shared"))
sys.path.insert(0, str(PROJECT_ROOT / "scripts" / "eval"))
import waf_guard  # noqa: E402
import verdict_extract as verdict  # noqa: E402
from browser_profile import REQUEST_HEADERS  # noqa: E402
from playwright_net import new_context as new_browser_context  # noqa: E402

DEFAULT_MUNICIPIOS_URL = "https://dados.tce.rs.gov.br/dados/auxiliar/municipios.csv"
UF_SIGLA = "RS"
UF_NOME = "Rio Grande do Sul"
GEMINI_BASE_URL = "https://generativelanguage.googleapis.com/v1beta"

LOGGER = logging.getLogger("fase2.cascade")
if not LOGGER.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(logging.Formatter("%(levelname)s %(name)s %(message)s"))
    LOGGER.addHandler(_handler)
LOGGER.propagate = False
LOGGER.setLevel(getattr(
    logging, os.environ.get("FASE2_LOG_LEVEL", "WARNING").upper(), logging.WARNING,
))

BAD_HOSTS = [
    "facebook.", "instagram.", "youtube.", "twitter.", "x.com",
    "linkedin.", "acheconcursos.", "pciconcursos.", "qconcursos.",
    "google.", "bing.", "duckduckgo.",
]

BUCKET_KEYWORDS = {
    "concursos": [
        "concurso", "concursos", "concursos publicos", "concurso publico",
    ],
    "processos": [
        "processo seletivo", "processos seletivos", "pss",
        "processos seletivos simplificados", "selecao publica",
        "selecoes publicas", "seletivo simplificado",
    ],
}

CONTAINER_KEYWORDS = [
    "editais", "edital", "publicacoes", "publicacao", "documentos",
    "transparencia", "contratacao", "oportunidades", "mural",
    "servicos", "portal",
]


# ---------------------------------------------------------------------------
# Text utilities
# ---------------------------------------------------------------------------
def norm(text: str) -> str:
    t = unicodedata.normalize("NFKD", text or "")
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = re.sub(r"[^a-z0-9\s]+", " ", t.lower())
    return re.sub(r"\s+", " ", t).strip()


def slugify(name: str) -> str:
    n = norm(name)
    n = re.sub(r"\b(da|de|do|das|dos)\b", "", n)
    return re.sub(r"[^a-z0-9]+", "", n)


def clean_url(url: str) -> str:
    url = (url or "").strip()
    if url and "://" not in url:
        url = "http://" + url
    return url.rstrip("/") if url else ""


# ---------------------------------------------------------------------------
# Page dataclass
# ---------------------------------------------------------------------------
@dataclass
class Page:
    url: str
    requested_url: str = ""
    status: int | None = 0
    title: str = ""
    text: str = ""
    html: str = ""
    links: list[tuple[str, str]] = field(default_factory=list)
    error: str = ""
    is_spa: bool = False  # served HTML is a JS shell (menu rendered client-side)
    is_antibot: bool = False  # served HTML is an anti-bot JS challenge (DDoS-Guard, etc.)

    @property
    def ok(self) -> bool:
        # A browser navigation may not expose a Response (SPA/client routing).
        # Missing status is therefore neutral; captured 4xx/5xx still fail.
        return not self.error and (
            self.status is None or 200 <= self.status < 400
        )


@dataclass(frozen=True)
class EvidenceSnapshot:
    """Deeply immutable evidence detached from its HTTP/browser producer."""

    html: str
    text: str
    title: str
    final_url: str
    requested_url: str = ""
    status: int | None = None
    source: str = "playwright"
    evidence_state: str = "renderizada"
    links: tuple[tuple[str, str], ...] = field(default_factory=tuple)

    def __post_init__(self) -> None:
        # Callers may still hand in Page.links (a list of lists/tuples). Detach it
        # now so later browser/page mutation cannot alter the adjudicated record.
        object.__setattr__(self, "links", tuple(
            (str(link[0]), str(link[1]))
            for link in (self.links or ()) if len(link) >= 2
        ))


# Compatibility name for the existing Tier 0/Tier 2 browser fallback API.
RenderedPage = EvidenceSnapshot


_DEFAULT_RENDERER = object()


# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------
def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(REQUEST_HEADERS)
    return s


# WAF / bot-block statuses where a plain-requests fetch is worth retrying with a
# browser TLS fingerprint (curl_cffi). 429/503 are rate/overload signals: do not
# immediately double the request there.
_FINGERPRINT_BLOCK_STATUSES = {403, 406, 409}
_RATE_LIMIT_STATUSES = {429, 503}


def _page_from_html(final_url: str, status: int | None, content_type: str,
                    html_text: str, requested_url: str = "") -> Page:
    if "text/html" not in content_type and "text/plain" not in content_type:
        return Page(
            url=final_url, requested_url=requested_url,
            status=status, error="not_html",
        )
    title = ""
    m = re.search(r"<title[^>]*>(.*?)</title>", html_text, re.I | re.S)
    if m:
        title = re.sub(r"\s+", " ", m.group(1)).strip()
    links = extract_links(final_url, html_text)
    body_text = extract_text(html_text)
    # SPA shell: framework markers present but the served HTML exposes almost no
    # links (the menu is rendered client-side). Tier 1 must render it to see it.
    html_low = html_text[:200000].lower()
    spa_markers = ("__next_data__" in html_low or "/_next/" in html_low
                   or "window.__nuxt__" in html_low or "data-reactroot" in html_low
                   or 'id="__nuxt"' in html_low)
    is_spa = spa_markers and len([h for h, _ in links if h.startswith("http")]) < 8
    # Anti-bot JS challenge (DDoS-Guard / "checking your browser"): a thin page
    # whose only job is to reload until a cookie is set. Not a real miss — flag
    # it so the report says "blocked", not "index not found".
    title_low = title.lower()
    challenge_title = ("one moment, please" in title_low or "just a moment" in title_low
                       or "checking your browser" in title_low
                       or "attention required" in title_low)
    # Explicit challenge markers (Cloudflare / DDoS-Guard) are unambiguous and
    # can ship a large body, so they flag regardless of page size. A generic
    # challenge title only flags when the page is thin (no real content).
    hard_markers = ("challenge-platform" in html_low or "/cdn-cgi/challenge" in html_low
                    or "_cf_chl_opt" in html_low or "cf_chl_" in html_low
                    or "ddos-guard" in html_low)
    is_antibot = hard_markers or (challenge_title and len(body_text) < 1500)
    return Page(url=final_url, requested_url=requested_url, status=status,
                title=title, html=html_text,
                text=body_text, links=links, is_spa=is_spa, is_antibot=is_antibot)


def _fetch_browser_impersonate(url: str, timeout: int) -> Page | None:
    """Fallback fetch with a real browser TLS fingerprint (curl_cffi).

    Returns None when curl_cffi is unavailable so the caller keeps the original
    requests-based result. Used only after a plain fetch is blocked/errors, so it
    never changes behaviour for sites that already work.
    """
    try:
        from curl_cffi import requests as creq
    except Exception:
        return None
    # curl_cffi does not read proxy env vars by default; pass them through so the
    # documented BR-proxy option also covers this fallback path.
    proxies = {}
    for scheme in ("http", "https"):
        val = os.environ.get(f"{scheme.upper()}_PROXY") or os.environ.get(f"{scheme}_proxy")
        if val:
            proxies[scheme] = val
    try:
        resp = creq.get(url, timeout=timeout, allow_redirects=True,
                        headers=REQUEST_HEADERS, impersonate="chrome",
                        proxies=proxies or None)
        return _page_from_html(
            str(resp.url), resp.status_code,
            resp.headers.get("content-type", ""), resp.text)
    except Exception as e:
        return Page(url=url, error=f"curl_cffi: {str(e)[:180]}")


def fetch_page(session: requests.Session, url: str, timeout: int = 15) -> Page:
    url = clean_url(url)
    if not url:
        return Page(url="", error="empty_url")
    if waf_guard.is_frozen(url):
        return Page(url=url, error="waf_frozen")
    try:
        resp = session.get(url, timeout=timeout, allow_redirects=True)
        if resp.status_code in _FINGERPRINT_BLOCK_STATUSES:
            alt = _fetch_browser_impersonate(url, timeout)
            if alt is not None and alt.ok:
                return alt
        elif resp.status_code in _RATE_LIMIT_STATUSES:
            pass
        page = _page_from_html(
            str(resp.url), resp.status_code,
            resp.headers.get("content-type", ""), resp.text,
            requested_url=url)
        return page
    except Exception as e:
        # Connection reset / TLS handshake rejected by a WAF — retry as a browser.
        alt = _fetch_browser_impersonate(url, timeout)
        if alt is not None and alt.ok:
            return alt
        return Page(url=url, error=str(e)[:200])


def extract_links(base_url: str, html: str) -> list[tuple[str, str]]:
    results = []
    seen = set()
    for m in re.finditer(r'<a\s[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', html, re.I | re.S):
        href_raw, link_text = m.group(1), m.group(2)
        link_text = re.sub(r"<[^>]+>", "", link_text).strip()
        link_text = re.sub(r"\s+", " ", link_text)
        try:
            href = urljoin(base_url, href_raw)
        except Exception:
            continue
        if href not in seen and href.startswith("http"):
            seen.add(href)
            results.append((href, link_text))
    return results


def extract_text(html: str) -> str:
    text = re.sub(r"<script[^>]*>.*?</script>", " ", html, flags=re.I | re.S)
    text = re.sub(r"<style[^>]*>.*?</style>", " ", text, flags=re.I | re.S)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"&[a-zA-Z]+;", " ", text)
    text = re.sub(r"&#\d+;", " ", text)
    return re.sub(r"\s+", " ", text).strip()


# ---------------------------------------------------------------------------
# Basic filters (cheap, deterministic, obvious rejections only)
# ---------------------------------------------------------------------------
SOFT_404_PATTERNS = [
    "pagina nao encontrada", "nao encontramos", "erro 404", "error 404",
    "not found", "pagina inexistente", "conteudo nao encontrado",
]


def is_soft_404(page: Page) -> bool:
    if not page.ok:
        return True
    blob = norm(page.title + " " + page.text[:500])
    return any(p in blob for p in SOFT_404_PATTERNS)


# Markers of a domain that RESOLVES but is not a live official site: an
# "under construction" stub, a parked domain, or a hosting placeholder. Such a
# page often still echoes the municipality name (so score_site_page would rate
# it highly) yet has no usable content. Accepting it as site_base silently
# blocks the grounded fallback that would find the real domain (frequently a
# .com.br the slug candidates never try). The markers are deliberately
# unambiguous to avoid rejecting a real site.
DEAD_SITE_PATTERNS = [
    "em construcao", "site em construcao", "pagina em construcao",
    "em manutencao", "site em manutencao",
    "hospedagem de site", "dominio gratis", "registre seu dominio",
    "compre este dominio", "domain for sale", "this domain is for sale",
]


def is_dead_site(page: Page) -> bool:
    """A reachable page that is a parked / hosting / under-construction stub."""
    blob = norm(page.title + " " + page.text[:800])
    return any(p in blob for p in DEAD_SITE_PATTERNS)


ANTIBOT_CHALLENGE_SIGNATURES = (
    "security checkpoint",
    "vercel security",
    "just a moment",
    "checking your browser",
    "attention required",
    "cloudflare",
    "verifying you are human",
    "enable javascript and cookies",
)


def is_antibot_challenge(page: Page) -> bool:
    """True only for an explicit browser/security checkpoint signature."""
    if page.error:
        # DNS, timeout, refused connection and other transport failures are not
        # browser challenges, even though they also produce a non-OK Page.
        return False
    blob = " ".join((page.title or "", page.text or "", page.html or "")).lower()
    return any(signature in blob for signature in ANTIBOT_CHALLENGE_SIGNATURES)


def is_broad_landing(url: str) -> bool:
    path = (urlparse(clean_url(url)).path or "/").strip("/").lower()
    return path in {"", "web", "home", "inicio", "index.php", "index.html",
                    "portal", "site"}


def is_pdf_or_file(url: str) -> bool:
    path = urlparse(url).path.lower()
    return any(path.endswith(ext) for ext in [".pdf", ".doc", ".docx", ".xls", ".xlsx"])


# A URL that points at ONE item (an individual edital, a single legislação, a
# single content/news page, a PDF) is never a stable index — the phase rules
# reject it. Even when Tier 3 picked it, it must not earn `confirmado`: the
# cascade keeps it `probable` and leaves promotion to the rendered AI verdict in
# the closing pass. None of the 24 golden index URLs match these patterns.
# Detalle INEQUIVOCO: nunca es un indice -> hard-block en el cierre.
HARD_DETAIL_PATTERNS = [
    r"/id/\d+",                # /concurso/id/200/
    r"/detalhe/",              # /detalhe/452/...
    r"/legislacao/detalhe",    # /legislacao/detalhe/3619/...
    r"[?&]slug=",              # ?slug=processo-seletivo (a single named item)
    r"\.pdf(\?|$)",
]
# AMBIGUO: /conteudo/N puede ser un DETALLE (una noticia/concurso unico) O el INDICE de
# la seccion — hay CMS que usan /conteudo/ID para sus paginas de LISTADO (p.ej. Imbe:
# "Concursos Publicos" vive en /conteudo/13400/... y "Processos Seletivos" en /conteudo/
# 13086/...). Por eso NO se hard-blockea: se juzga por contenido (render + item + IA).
AMBIGUOUS_DETAIL_PATTERNS = [
    r"/conteudos?/\d+",        # /site/conteudos/5848-... o /conteudo/13400 (indice o detalle)
]
DETAIL_URL_PATTERNS = HARD_DETAIL_PATTERNS + AMBIGUOUS_DETAIL_PATTERNS


def is_detail_url(url: str) -> bool:
    """Union (duro + ambiguo). Usado por el cascade para bajar la confianza a probable."""
    u = (url or "").lower()
    return any(re.search(p, u) for p in DETAIL_URL_PATTERNS)


def is_hard_detail_url(url: str) -> bool:
    """Solo detalle INEQUIVOCO (/id/N, /detalhe/, .pdf, ?slug=). El cierre hard-blockea
    estas; deja pasar /conteudo/N (ambiguo) a verificacion por contenido."""
    u = (url or "").lower()
    return any(re.search(p, u) for p in HARD_DETAIL_PATTERNS)


# ---------------------------------------------------------------------------
# Candidate: a URL with metadata
# ---------------------------------------------------------------------------
@dataclass
class Candidate:
    url: str
    source: str  # "menu_link", "container_link", "grounding", "playwright"
    menu_text: str = ""
    page: Page | None = None
    accessible: bool = True
    evidence_state: str = "completa"
    source_kind: str = "desconocido"
    authority: str = "desconocida"
    identity: str = "desconocida"
    page_role: str = "desconocido"
    decision: str = "revisar"
    note: str = ""
    provenance: list[dict] = field(default_factory=list)
    bucket: str = "combinado"
    content_preview: str = ""
    evidence_snapshot: EvidenceSnapshot | None = field(default=None, repr=False)
    source_tier: str = ""
    record: CandidateRecord | None = field(default=None, repr=False)
    # Discovery hint only. It records which bucket led us to the candidate so
    # an uncertain adjudication can be returned to that bucket for review.
    # It never decides the final bucket; forma/tipo from page content do that.
    bucket_hint: str = ""

    def __post_init__(self) -> None:
        if not self.bucket_hint:
            self.bucket_hint = _discovery_bucket_hint(self.menu_text, self.url)

    @property
    def fetchable(self) -> bool:
        """Compatibility alias; fetchability now means operational accessibility."""
        return self.accessible

    @fetchable.setter
    def fetchable(self, value: bool) -> None:
        self.accessible = bool(value)
        if not self.accessible:
            self.evidence_state = "error_fetch"

    @property
    def eligible(self) -> bool:
        return self.decision in {
            "indice_oficial", "indice_oficial_combinado", "portal_externo_oficial",
        }

    @property
    def candidate_id(self) -> str:
        return self.record.candidate_id if self.record else ""


def _candidate_url_key(url: str) -> str:
    """Normalize a URL with the repository's existing comparison semantics."""
    parsed = urlparse(clean_url(url))
    host = parsed.netloc.lower()
    if host.startswith("www."):
        host = host[4:]
    path = (parsed.path or "").rstrip("/")
    query = "&".join(
        f"{key.lower()}={value}" for key, value in
        sorted(parse_qsl(parsed.query, keep_blank_values=True))
    )
    base = f"{host}{path}"
    return f"{base}?{query}" if query else base


def _freeze_provenance(
        provenance: list[dict] | tuple[dict, ...],
        ) -> tuple[tuple[tuple[str, str], ...], ...]:
    """Detach arbitrary provenance mappings into deterministic immutable pairs."""
    frozen = []
    for item in provenance or ():
        if not isinstance(item, dict):
            continue
        frozen.append(tuple(sorted(
            (str(key), json.dumps(value, ensure_ascii=False, sort_keys=True))
            for key, value in item.items()
        )))
    return tuple(frozen)


@dataclass(frozen=True)
class CandidateRecord:
    """Immutable, fully adjudicated candidate built before Tier 3 selection.

    ``candidate_id`` is ``v1:`` plus SHA-1 over a deterministic JSON payload
    containing the normalized final URL (lower-case host, no fragment, no
    trailing slash, sorted query using the repository URL normalizer), source,
    tier, normalized municipality, adjudicated bucket and a SHA-1 fingerprint
    of the complete immutable EvidenceSnapshot. Including the snapshot digest
    distinguishes different captures of the same URL. The ID is only an audit
    and serialization key; a CandidateRecord must never be reconstructed from it.
    """

    candidate_id: str
    requested_url: str
    final_url: str
    source: str
    tier: str
    municipio: str
    bucket_hint: str
    evidence_snapshot: EvidenceSnapshot = field(repr=False)
    authority: str = "desconocida"
    identity: str = "desconocida"
    page_role: str = "desconocido"
    evidence_state: str = "error_fetch"
    bucket: str = "combinado"
    decision: str = "revisar"
    reason: str = "sin razon"
    source_kind: str = "desconocido"
    accessible: bool = False
    provenance: tuple[tuple[tuple[str, str], ...], ...] = field(
        default_factory=tuple, repr=False,
    )
    menu_text: str = ""

    @property
    def eligible(self) -> bool:
        return self.decision in {
            "indice_oficial", "indice_oficial_combinado", "portal_externo_oficial",
        }

    @property
    def url(self) -> str:
        return self.final_url

    @property
    def source_tier(self) -> str:
        return self.tier

    @property
    def note(self) -> str:
        return self.reason

    @property
    def fetchable(self) -> bool:
        return self.accessible

    @property
    def page(self) -> Page:
        snapshot = self.evidence_snapshot
        page = _page_from_html(
            snapshot.final_url, snapshot.status, "text/html; charset=UTF-8",
            snapshot.html, requested_url=snapshot.requested_url,
        )
        page.title = snapshot.title or page.title
        page.text = snapshot.text or page.text
        page.links = list(snapshot.links)
        return page

    @property
    def content_preview(self) -> str:
        return self.evidence_snapshot.text[:1200]


@dataclass(frozen=True)
class SelectedResource:
    """Exact CandidateRecord instance selected for one canonical bucket."""

    bucket: str
    candidate: CandidateRecord
    reason: str = "selector eligio candidate_id existente y elegible"


@dataclass(frozen=True)
class FinalDecision:
    """CSV-facing state derived only from SelectedResource and the contract."""

    bucket: str
    status: str
    decision: str
    url: str
    candidate_id: str
    reason: str


@dataclass(frozen=True)
class BucketCandidateEvidence:
    """Independent bucket association for one selected candidate and evidence."""

    bucket: str
    candidate: Candidate | CandidateRecord
    snapshot: EvidenceSnapshot


def _discovery_bucket_hint(menu_text: str, url: str) -> str:
    """Infer only the discovery origin; never the candidate's final bucket."""
    discovery_text = norm(f"{menu_text} {unquote(urlparse(url).path)}")
    matched = {
        bucket for bucket, keywords in BUCKET_KEYWORDS.items()
        if any(keyword in discovery_text for keyword in keywords)
    }
    if len(matched) == 1:
        return next(iter(matched))
    if len(matched) > 1:
        return "ambos"
    return ""


# ---------------------------------------------------------------------------
# TIER 0: Find/confirm site base
# ---------------------------------------------------------------------------
def domain_candidates(municipio: str) -> list[str]:
    # Only the safe, full-name slugs (no collision-prone heuristics like
    # first-word or pm+initials). When these miss, Tier 2 grounded search
    # discovers the real domain — see tier2_find_site_grounded.
    slugs = []
    s1 = slugify(municipio)
    s2 = re.sub(r"[^a-z0-9]+", "", norm(municipio))
    for s in [s1, s2]:
        if s and s not in slugs:
            slugs.append(s)
    urls = []
    for slug in slugs:
        for prefix in ["www.", "", "pm"]:
            host = f"{prefix}{slug}.rs.gov.br"
            urls.extend([f"https://{host}/", f"http://{host}/"])
        urls.append(f"https://{slug}.atende.net/")
    seen = set()
    return [u for u in urls if u not in seen and not seen.add(u)]


def score_site_page(page: Page, municipio: str) -> int:
    """How strongly a page looks like the official prefeitura homepage.

    Used both by the free slug discovery (Tier 0) and the grounded domain
    discovery fallback, so the validation bar is identical regardless of how
    the URL was found.
    """
    # A parked / hosting / under-construction stub is never a usable site_base,
    # even if it echoes the municipality name. Reject it hard (negative score) so
    # Tier 0 misses and the grounded site discovery runs to find the real domain.
    if is_soft_404(page) or is_dead_site(page):
        return -100
    blob = norm(page.title + " " + page.text[:2000])
    muni_norm = norm(municipio)
    score = 0
    if muni_norm in blob:
        score += 10
    if "prefeitura" in blob:
        score += 5
    if ".rs.gov.br" in page.url:
        score += 3
    if ".atende.net" in page.url:
        score += 2
    return score


def is_matching_official_municipality_domain(page: Page, municipio: str) -> bool:
    """Confirm an official RS host whose municipal label matches the municipality.

    Some official homes return a generic client-rendered shell, so their static
    HTML has neither the municipality name nor ``prefeitura``. Domain confirmation
    must not confuse that missing rendered content with a missing site.
    """
    if is_soft_404(page) or is_dead_site(page):
        return False
    host = (urlparse(page.url).hostname or "").lower().rstrip(".")
    suffix = ".rs.gov.br"
    if not host.endswith(suffix):
        return False
    municipal_labels = host[:-len(suffix)].split(".")
    return slugify(municipio) in municipal_labels


def _official_host(url: str, municipio: str) -> str:
    """Return the host only when the existing Tier 0 identity check accepts it."""
    host = (urlparse(clean_url(url)).hostname or "").lower().rstrip(".")
    neutral = Page(url=clean_url(url), status=200)
    return host if is_matching_official_municipality_domain(neutral, municipio) else ""


def _candidate_identity_matches(page: Page, municipio: str) -> bool:
    """Validate municipal identity from the final URL and rendered content.

    Official RS municipality hosts use the existing hostname/slug contract. A
    delegated external portal must instead identify the target municipality in
    its rendered title/body; a generic third party is not accepted by URL alone.
    """
    host = (urlparse(page.url).hostname or "").lower().rstrip(".")
    if host.endswith(".rs.gov.br"):
        return is_matching_official_municipality_domain(page, municipio)
    municipality = norm(municipio)
    identity_blob = norm(f"{page.title}\n{page.text[:3000]}")
    return bool(municipality and municipality in identity_blob)


def _provenance_confirms(provenance: list[dict] | tuple[dict, ...],
                         municipio: str) -> bool:
    """Confirm a delegated chain from explicit official navigation evidence."""
    target = norm(municipio)
    for item in provenance or ():
        if not isinstance(item, dict):
            continue
        if item.get("kind") not in {
            "official_navigation", "official_referrer", "official_brand",
        }:
            continue
        if norm(str(item.get("municipio", ""))) == target:
            return True
    return False


def _candidate_identity_state(page: Page, municipio: str,
                              provenance: list[dict] | tuple[dict, ...]) -> str:
    if is_soft_404(page) or is_dead_site(page):
        return "rechazada"
    host = (urlparse(page.url).hostname or "").lower().rstrip(".")
    if host.endswith(".rs.gov.br"):
        municipal_labels = host[:-len(".rs.gov.br")].split(".")
        if slugify(municipio) not in municipal_labels:
            return "rechazada"
    target = norm(municipio)
    identity_blob = norm(f"{page.title}\n{page.text[:3000]}")
    if target and target in identity_blob:
        return "confirmada"
    if _provenance_confirms(provenance, municipio):
        return "confirmada"
    return "desconocida"


def _candidate_source_and_authority(
        page: Page, municipio: str,
        provenance: list[dict] | tuple[dict, ...], identity: str,
        source: str = "") -> tuple[str, str]:
    """Derive provenance independently of page structure; never from URL slug."""
    host = (urlparse(page.url).hostname or "").lower().rstrip(".")
    official_chain = _provenance_confirms(provenance, municipio)
    explicitly_external = source == "portal_externo_delegado"
    if (official_chain or explicitly_external) and not host.endswith(".rs.gov.br"):
        if not official_chain:
            return "portal_externo_delegado", "desconocida"
        return "portal_externo_delegado", "confirmada"
    if host.endswith(".rs.gov.br"):
        official_brand = bool(re.search(
            r"\b(?:prefeitura|municipio|portal oficial)\b",
            norm(f"{page.title}\n{page.text[:3000]}"),
        ))
        return (
            "dominio_oficial_prefeitura",
            "confirmada" if identity == "confirmada" and official_brand else "desconocida",
        )
    if official_chain:
        return "portal_externo_delegado", "confirmada"
    return "desconocido", "desconocida"


def _snapshot_fingerprint(snapshot: EvidenceSnapshot) -> str:
    payload = {
        "html": snapshot.html,
        "text": snapshot.text,
        "title": snapshot.title,
        "final_url": _candidate_url_key(snapshot.final_url),
        "requested_url": snapshot.requested_url,
        "status": snapshot.status,
        "source": snapshot.source,
        "evidence_state": snapshot.evidence_state,
        "links": snapshot.links,
    }
    raw = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha1(raw).hexdigest()


def _candidate_record_id(*, final_url: str, source: str, tier: str,
                         municipio: str, bucket: str,
                         snapshot: EvidenceSnapshot) -> str:
    payload = {
        "url": _candidate_url_key(final_url),
        "source": source,
        "tier": tier,
        "municipio": norm(municipio),
        "bucket": bucket,
        "snapshot": _snapshot_fingerprint(snapshot),
    }
    raw = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
    ).encode("utf-8")
    return "v1:" + hashlib.sha1(raw).hexdigest()


def build_candidate_record(
        *, requested_url: str, source: str, tier: str, municipio: str,
        bucket_hint: str, evidence: EvidenceSnapshot,
        menu_text: str = "",
        provenance: list[dict] | tuple[dict, ...] = (),
        ) -> CandidateRecord:
    """Run the one deterministic adjudicator over an already captured snapshot.

    This function never performs I/O. All authority, identity, structure,
    evidence, bucket, decision and reason dimensions are fixed here exactly once.
    """
    requested = evidence.requested_url or requested_url
    final_url = clean_url(evidence.final_url or requested)
    page = _page_from_html(
        final_url, evidence.status, "text/html; charset=UTF-8",
        evidence.html, requested_url=requested,
    )
    page.title = evidence.title or page.title
    page.text = evidence.text or page.text
    page.links = list(evidence.links) or page.links
    evidence_state = (
        "error_fetch" if evidence.status is not None and evidence.status >= 400
        else "incompleta_antibot" if is_antibot_challenge(page)
        else evidence.evidence_state
    )
    accessible = evidence_state != "error_fetch"
    identity = _candidate_identity_state(page, municipio, provenance)
    source_kind, authority = _candidate_source_and_authority(
        page, municipio, provenance, identity, source,
    )
    anchors = [{"href": href, "text": text} for href, text in page.links]
    contracts = {
        bucket: verdict.evaluate_candidate_contract(
            page.text, bucket, title=page.title, anchors=anchors,
            source_kind=source_kind, authority=authority, identity=identity,
            evidence_state=evidence_state, accessible=accessible,
            provenance=provenance,
        )
        for bucket in ("concursos", "processos")
    }
    accepted = [contract for contract in contracts.values() if contract.decision in {
        "indice_oficial", "indice_oficial_combinado", "portal_externo_oficial",
    }]
    hint_key = (
        "concursos" if bucket_hint == "concursos" else
        "processos" if bucket_hint == "processos" else ""
    )
    if accepted:
        diagnostic = next((
            contract for contract in accepted
            if contract.bucket == "combinado"
        ), None) or (
            contracts.get(hint_key) if contracts.get(hint_key) in accepted else accepted[0]
        )
    elif hint_key:
        diagnostic = contracts[hint_key]
    else:
        diagnostic = next(
            (contract for contract in contracts.values()
             if contract.page_role != "desconocido"),
            contracts["concursos"],
        )
    reason = diagnostic.note or f"decision={diagnostic.decision} sin detalle"
    candidate_id = _candidate_record_id(
        final_url=final_url, source=source, tier=tier,
        municipio=municipio, bucket=diagnostic.bucket, snapshot=evidence,
    )
    record = CandidateRecord(
        candidate_id=candidate_id,
        requested_url=requested,
        final_url=final_url,
        source=source,
        tier=tier,
        municipio=municipio,
        bucket_hint=bucket_hint,
        evidence_snapshot=evidence,
        authority=diagnostic.authority,
        identity=diagnostic.identity,
        page_role=diagnostic.page_role,
        evidence_state=diagnostic.evidence_state,
        bucket=diagnostic.bucket,
        decision=diagnostic.decision,
        reason=reason,
        source_kind=diagnostic.source_kind,
        accessible=diagnostic.accessible,
        provenance=_freeze_provenance(provenance),
        menu_text=menu_text,
    )
    LOGGER.info("candidate_record %s", json.dumps({
        "candidate_id": record.candidate_id,
        "municipio": record.municipio,
        "source": record.source,
        "tier": record.tier,
        "requested_url": record.requested_url,
        "final_url": record.final_url,
        "authority": record.authority,
        "identity": record.identity,
        "page_role": record.page_role,
        "evidence_state": record.evidence_state,
        "bucket": record.bucket,
        "decision": record.decision,
        "reason": record.reason,
    }, ensure_ascii=False, sort_keys=True))
    return record


def candidate_from_evidence(
        requested_url: str, source: str, menu_text: str, municipio: str,
        evidence: Page | EvidenceSnapshot,
        provenance: list[dict] | tuple[dict, ...] = (), *,
        tier: str = "", bucket_hint: str = "") -> Candidate:
    """Compatibility adapter around the canonical CandidateRecord adjudicator."""
    snapshot = (
        _snapshot_with_requested_url(evidence, requested_url)
        if isinstance(evidence, EvidenceSnapshot)
        else _snapshot_from_page(requested_url, evidence)
    )
    record = build_candidate_record(
        requested_url=requested_url, source=source, tier=tier,
        municipio=municipio, bucket_hint=bucket_hint,
        evidence=snapshot, menu_text=menu_text, provenance=provenance,
    )
    page = record.page
    return Candidate(
        url=record.final_url, source=source, menu_text=menu_text,
        page=page, content_preview=record.content_preview,
        evidence_snapshot=snapshot, accessible=record.accessible,
        evidence_state=record.evidence_state, source_kind=record.source_kind,
        authority=record.authority, identity=record.identity,
        page_role=record.page_role, decision=record.decision,
        note=record.reason, provenance=[dict(item) for item in provenance],
        bucket=record.bucket, source_tier=tier, bucket_hint=bucket_hint,
        record=record,
    )


def _snapshot_from_page(requested_url: str, page: Page) -> EvidenceSnapshot:
    """Detach a complete immutable snapshot from a static HTTP response."""
    evidence_state = (
        "error_fetch" if page.error or (
            page.status is not None and page.status >= 400
        ) else
        "incompleta_antibot" if is_antibot_challenge(page) else
        "completa"
    )
    return EvidenceSnapshot(
        html=page.html or "", text=page.text or "", title=page.title or "",
        requested_url=page.requested_url or requested_url,
        final_url=page.url or requested_url, status=page.status,
        source="requests", evidence_state=evidence_state,
        links=tuple(page.links),
    )


def _snapshot_with_requested_url(snapshot: EvidenceSnapshot,
                                 requested_url: str) -> EvidenceSnapshot:
    """Defensively complete legacy snapshots without mutating shared evidence."""
    if snapshot.requested_url:
        return snapshot
    return EvidenceSnapshot(
        html=snapshot.html, text=snapshot.text, title=snapshot.title,
        requested_url=requested_url, final_url=snapshot.final_url,
        status=snapshot.status, source=snapshot.source,
        evidence_state=snapshot.evidence_state,
        links=snapshot.links,
    )


def hydrate_candidate(
        candidate: Candidate, municipio: str, *, session=None, timeout: int = 15,
        evidence: Page | EvidenceSnapshot | None = None,
        official_url: str = "", render_page=_DEFAULT_RENDERER) -> Candidate:
    """Canonical hydration path for every discovery producer.

    Discovery metadata never stands in for content. The function obtains or
    accepts real evidence, detaches it into one immutable snapshot, and then
    traverses the existing deterministic candidate contract.
    """
    requested_url = candidate.evidence_snapshot.requested_url if (
        candidate.evidence_snapshot and candidate.evidence_snapshot.requested_url
    ) else candidate.url
    if evidence is None:
        if session is None:
            raise ValueError("session is required when evidence is not supplied")
        if official_url:
            evidence = fetch_page_with_official_fallback(
                session, requested_url, municipio, official_url, timeout,
                render_page, preserve_snapshot=True,
            )
        else:
            evidence = fetch_page(session, requested_url, timeout)
    snapshot = (
        _snapshot_with_requested_url(evidence, requested_url)
        if isinstance(evidence, EvidenceSnapshot)
        else _snapshot_from_page(requested_url, evidence)
    )
    evaluated = candidate_from_evidence(
        requested_url, candidate.source, candidate.menu_text, municipio, snapshot,
        provenance=tuple(dict(item) for item in candidate.provenance),
        tier=candidate.source_tier, bucket_hint=candidate.bucket_hint,
    )
    return evaluated


def _apply_candidate_evidence(candidate: Candidate, evidence: Page | EvidenceSnapshot,
                              municipio: str) -> Candidate:
    """Mutate an existing discovery candidate from already captured evidence."""
    evaluated = hydrate_candidate(candidate, municipio, evidence=evidence)
    for name in (
        "url", "source", "page", "accessible", "evidence_state", "source_kind",
        "authority", "identity", "page_role", "decision", "note", "provenance",
        "bucket", "content_preview", "evidence_snapshot", "source_tier",
        "record",
    ):
        setattr(candidate, name, getattr(evaluated, name))
    return candidate


def fetch_page_with_official_fallback(
        session: requests.Session, url: str, municipio: str,
        official_url: str, timeout: int = 15,
        render_page=_DEFAULT_RENDERER,
        preserve_snapshot: bool = False) -> Page | EvidenceSnapshot:
    """Fetch an unequivocal municipal official URL, rendering only a bad shell.

    The normal fetch remains the source of the fallback decision. Playwright is
    attempted once only when the requested/final host is exactly the official
    host accepted by Tier 0's existing municipality identity mechanism.
    """
    # Preserve the caller's exact URL for the browser attempt and audit field;
    # fetch_page performs its own established URL cleanup for the normal path.
    requested_url = (url or "").strip()
    normal_page = fetch_page(session, requested_url, timeout)

    expected_host = _official_host(official_url, municipio)
    response_host = _official_host(normal_page.url, municipio)
    same_exact_official_host = bool(
        expected_host and response_host and expected_host == response_host
    )
    recognized_challenge = is_antibot_challenge(normal_page)
    unverifiable_shell = normal_page.ok and (
        is_soft_404(normal_page) or is_dead_site(normal_page)
    )
    if not (same_exact_official_host
            and (recognized_challenge or unverifiable_shell)):
        return normal_page
    if render_page is None:
        return normal_page
    if render_page is _DEFAULT_RENDERER:
        render_page = render_page_sync

    try:
        # One direct render attempt. The renderer never calls this wrapper, so
        # failure cannot recurse into another Playwright attempt.
        rendered = render_page(requested_url)
    except Exception as e:
        print(f"      official fallback render error: {e}", flush=True)
        return normal_page
    if not rendered:
        return normal_page

    final_host = _official_host(rendered.final_url, municipio)
    if not final_host or final_host != expected_host:
        return normal_page

    recovered = _page_from_html(
        rendered.final_url, rendered.status, "text/html; charset=UTF-8",
        rendered.html, requested_url=requested_url,
    )
    # Keep the browser's DOM-derived fields faithfully even for minimal test
    # doubles. `ok`, soft-404 and dead-site remain normal Page validations.
    recovered.title = rendered.title or recovered.title
    recovered.text = rendered.text or recovered.text
    if (not recovered.ok or is_antibot_challenge(recovered)
            or is_soft_404(recovered) or is_dead_site(recovered)):
        return normal_page
    if not is_matching_official_municipality_domain(recovered, municipio):
        return normal_page
    if norm(municipio) not in norm(f"{recovered.title} {recovered.text}"):
        return normal_page

    # fetch_page has no page cache: only WAF host-state is cached. Therefore no
    # checkpoint is promoted and no rendered DOM needs a cache replacement.
    if preserve_snapshot:
        return _snapshot_with_requested_url(rendered, requested_url)
    return recovered


def tier0_find_site(session: requests.Session, municipio: str,
                    timeout: int = 15,
                    render_page=_DEFAULT_RENDERER) -> Page | None:
    candidates = domain_candidates(municipio)
    best = None
    best_score = -1
    for url in candidates:
        page = fetch_page_with_official_fallback(
            session, url, municipio, url, timeout, render_page,
        )
        if is_antibot_challenge(page):
            continue
        if not page.ok:
            continue
        score = score_site_page(page, municipio)
        if score > best_score:
            best_score = score
            best = page
    if best and (best_score >= 5
                 or is_matching_official_municipality_domain(best, municipio)):
        migrated = _check_migration(session, best, timeout)
        if migrated:
            return migrated
        return best
    return best if best and best_score > 0 else None


def _check_migration(session: requests.Session, page: Page,
                     timeout: int = 15) -> Page | None:
    migration_patterns = [
        r"novo\s+site", r"novo\s+endereco", r"novo\s+portal",
        r"mudou\s+para", r"acesse\s+o\s+novo",
    ]
    blob = norm(page.text[:3000])
    if not any(re.search(p, blob) for p in migration_patterns):
        return None
    for href, text in page.links:
        text_n = norm(text)
        if any(re.search(p, text_n) for p in migration_patterns) or "novo site" in text_n:
            new_page = fetch_page(session, href, timeout)
            if new_page.ok and not is_soft_404(new_page):
                return new_page
    return None


# ---------------------------------------------------------------------------
# TIER 1: Free link discovery
# ---------------------------------------------------------------------------
# Recurring "where the index lives" CMS paths, confirmed by hand across many RS
# municipalities (see project_patrones_indices_codeables). Probing them, derived
# only from the site HOST (no hardcoded municipality name/IP/portal), widens the
# candidate set toward the canonical index when the menu surfaced a worse or
# year-filtered URL — e.g. a govbr CMS exposes the combined index at
# `/site/concursos` even when the menu only links a single `?tipo=N` filter.
# These are CANDIDATES only: Tier 3's discrete verifier is still the only thing
# that confirms a bucket, so probing cannot create false positives beyond Tier 3.
# Only HIGH-PRECISION, low-collision CMS signatures: paths that render a
# distinctive combined index ("Concursos e Seleções Públicas" with a Tipo field,
# "Editais de Concursos" with a Categoria dropdown, the `/portal/editais/N`
# template). Generic paths like bare `/concursos` or `/portal-da-transparencia/
# concursos-publicos` were intentionally DROPPED: they collide on big-city sites
# (caused golden F-POS/WRNG on Aceguá, Porto Alegre, São Leopoldo) and are
# already found by the menu/link discovery anyway, so probing them only adds risk.
PROBE_PATHS_DEFAULT = [
    "/site/concursos",
    "/concurso",
    "/portal/editais/3",
]
PROBE_PATHS_ATENDE = [
    "/cidadao/pagina/concursos",
    "/cidadao/pagina/processos-seletivos",
    "/cidadao/pagina/concurso-e-processos-seletivos",
]


# A probed path frequently resolves to a soft-404 or a generic CMS fallback
# (e.g. atende serving "Páginas" / "Valores de Diárias", or a "Não Encontrado"
# stub returned with HTTP 200). Those must NOT reach Tier 3, which occasionally
# over-confirms them. So a probe page is only accepted if it (a) isn't a stub and
# (b) actually talks about concursos / processos seletivos. This is a content
# gate, not a scorer — it just keeps the probe from inventing candidates.
PROBE_REJECT_TITLE = [
    "nao encontrado", "nao encontrada", "acesso negado", "forbidden",
    "pagina inexistente", "erro 404", "error 404", "indisponivel",
]
PROBE_RELEVANT_KEYWORDS = [
    "concurso", "processo seletivo", "processos seletivos",
    "selecao publica", "selecoes publicas", "seletivo simplificado",
]


def _probe_page_is_index_like(page: Page) -> bool:
    """A probed page is a usable candidate only if it is a real, on-topic page
    (not a stub/fallback) that mentions concursos or processos seletivos."""
    if not page.ok or is_soft_404(page) or is_dead_site(page):
        return False
    if is_broad_landing(page.url):
        return False
    title_n = norm(page.title)
    if any(p in title_n for p in PROBE_REJECT_TITLE):
        return False
    blob = norm(page.text[:3000])
    return any(k in blob for k in PROBE_RELEVANT_KEYWORDS)


def _probe_known_index_paths(session: requests.Session, home: Page,
                             seen_urls: set[str], municipio: str,
                             timeout: int = 15) -> list[Candidate]:
    """Probe well-known CMS index paths derived from the site host.

    Returns fetched Candidates (source='probe') for paths that resolve to a real,
    on-topic page (see _probe_page_is_index_like) not already discovered.
    Platform-selected by host so the cost stays small (~3-6 cheap GETs). Pure
    requests, no AI, no scoring — Tier 3 still decides which (if any) to confirm.
    """
    parsed = urlparse(home.url)
    host = parsed.netloc.lower()
    if not host:
        return []
    base = f"{parsed.scheme}://{parsed.netloc}"
    paths = PROBE_PATHS_ATENDE if host.endswith("atende.net") else PROBE_PATHS_DEFAULT
    probes: list[Candidate] = []
    for path in paths:
        url = base + path
        if url in seen_urls or clean_url(url) in seen_urls:
            continue
        page = fetch_page(session, url, min(timeout, 10))
        if not _probe_page_is_index_like(page):
            continue
        seen_urls.add(url)
        candidate = Candidate(
            url=url, source="probe", menu_text="(probe: known index path)",
            source_tier="tier1",
        )
        probes.append(hydrate_candidate(candidate, municipio, evidence=page))
    return probes


def _official_navigation_provenance(home: Page, municipio: str,
                                    label: str) -> list[dict]:
    """Represent an official click chain only when the home content proves it."""
    blob = norm(f"{home.title}\n{home.text[:3000]}")
    if norm(municipio) not in blob or not re.search(
            r"\b(?:prefeitura|municipio|portal oficial)\b", blob):
        return []
    return [{
        "kind": "official_navigation", "municipio": municipio,
        "referrer": home.url, "label": label,
    }]


def tier1_collect_candidates(session: requests.Session, home: Page,
                             municipio: str, timeout: int = 15,
                             render_page=_DEFAULT_RENDERER) -> list[Candidate]:
    """Scan home page links and one level of container pages for relevant URLs."""
    candidates = []
    seen_urls: set[str] = set()
    all_keywords = []
    for kws in BUCKET_KEYWORDS.values():
        all_keywords.extend(kws)

    # Direct links from home page
    for href, link_text in home.links:
        host = urlparse(href).netloc.lower()
        if any(bad in host for bad in BAD_HOSTS):
            continue
        if is_pdf_or_file(href) or is_broad_landing(href):
            continue
        text_n = norm(link_text)
        href_n = norm(unquote(urlparse(href).path))
        if any(kw in text_n or kw in href_n for kw in all_keywords):
            if href not in seen_urls:
                seen_urls.add(href)
                candidates.append(Candidate(
                    url=href, source="menu_link", menu_text=link_text,
                    source_tier="tier1",
                    provenance=_official_navigation_provenance(home, municipio, link_text),
                ))

    # One level deep: follow container-like links
    # Skip links already captured as bucket candidates
    bucket_hrefs = {c.url for c in candidates}
    container_urls = []
    for href, link_text in home.links:
        if href in bucket_hrefs or href in seen_urls:
            continue
        text_n = norm(link_text)
        href_n = norm(unquote(urlparse(href).path))
        if any(kw in text_n or kw in href_n for kw in CONTAINER_KEYWORDS):
            if not is_broad_landing(href) and not is_pdf_or_file(href):
                seen_urls.add(href)
                container_urls.append((href, link_text))

    for container_href, container_text in container_urls[:4]:
        container_page = fetch_page(session, container_href, min(timeout, 10))
        if not container_page.ok or is_soft_404(container_page):
            continue
        for href, link_text in container_page.links:
            host = urlparse(href).netloc.lower()
            if any(bad in host for bad in BAD_HOSTS):
                continue
            if is_pdf_or_file(href) or is_broad_landing(href):
                continue
            text_n = norm(link_text)
            href_n = norm(unquote(urlparse(href).path))
            if any(kw in text_n or kw in href_n for kw in all_keywords):
                if href not in seen_urls:
                    seen_urls.add(href)
                    candidates.append(Candidate(
                        url=href, source="container_link",
                        source_tier="tier1",
                        menu_text=f"{container_text} > {link_text}",
                        provenance=_official_navigation_provenance(
                            home, municipio, f"{container_text} > {link_text}",
                        ),
                    ))

    # Fetch each candidate page to get content for Tier 3
    candidates = [
        hydrate_candidate(
            c, municipio, session=session, timeout=min(timeout, 10),
            official_url=home.url, render_page=render_page,
        )
        for c in candidates
    ]

    # Drill-down: a bucket parent page (e.g. /concurso) often links to more
    # specific sub-indexes (e.g. /concurso/categoria/25/concurso). Also follows
    # same-level siblings (same path depth, different leaf) so that a concursos
    # page can lead to the processos page beside it.
    drill: list[Candidate] = []
    for c in [c for c in candidates if c.fetchable and c.page]:
        parent_path = urlparse(c.url).path.rstrip("/")
        parent_host = urlparse(c.url).netloc.lower()
        parent_parent = "/".join(parent_path.split("/")[:-1]) if "/" in parent_path.lstrip("/") else ""
        for href, link_text in c.page.links:
            if href in seen_urls or len(drill) >= 12:
                continue
            pu = urlparse(href)
            if pu.netloc.lower() != parent_host:
                continue
            child_path = pu.path.rstrip("/")
            is_child = child_path.startswith(parent_path + "/")
            is_sibling = (parent_parent
                          and child_path.startswith(parent_parent + "/")
                          and child_path != parent_path
                          and child_path.count("/") == parent_path.count("/"))
            if not is_child and not is_sibling:
                continue
            if is_pdf_or_file(href) or is_broad_landing(href):
                continue
            text_n = norm(link_text)
            href_n = norm(unquote(pu.path))
            if any(kw in text_n or kw in href_n for kw in all_keywords):
                seen_urls.add(href)
                drill.append(Candidate(
                    url=href, source="drilldown",
                    source_tier="tier1",
                    menu_text=f"{c.menu_text} > {link_text}",
                    provenance=list(c.provenance),
                ))
    drill = [
        hydrate_candidate(
            c, municipio, session=session, timeout=min(timeout, 10),
            official_url=home.url, render_page=render_page,
        )
        for c in drill
    ]
    candidates.extend(drill)

    # Parameter normalization: if a candidate has ano=YYYY, add ano=0 variant
    # (all years) so Tier 3 can pick the canonical unfiltered view.
    param_variants: list[Candidate] = []
    for c in [c for c in candidates if c.fetchable]:
        if "ano=" in c.url and not re.search(r"[?&]ano=0(?:&|$)", c.url):
            variant_url = re.sub(r"([?&]ano=)\d{4}", r"\g<1>0", c.url)
            if variant_url not in seen_urls:
                seen_urls.add(variant_url)
                param_variants.append(Candidate(
                    url=variant_url, source="param_variant",
                    source_tier="tier1",
                    menu_text=f"{c.menu_text} (all years)",
                    provenance=list(c.provenance),
                ))
    param_variants = [
        hydrate_candidate(
            c, municipio, session=session, timeout=min(timeout, 10),
            official_url=home.url, render_page=render_page,
        )
        for c in param_variants
    ]
    candidates.extend(param_variants)

    return candidates


# ---------------------------------------------------------------------------
# TIER 2: Gemini grounded search
# ---------------------------------------------------------------------------
def gemini_api_key() -> str:
    return os.environ.get("GEMINI_API_KEY", "")


GEMINI_POST_CALLS = 0
GEMINI_POST_CALLS_FREE = 0
GEMINI_POST_CALLS_PAID = 0

_GEMINI_FREE_WINDOW: list[tuple[float, int]] = []
_GEMINI_FREE_COOLDOWN_UNTIL = 0.0
_GEMINI_FREE_COOLDOWN_STREAK = 0
_GEMINI_FREE_DISABLED_FOR_RUN = False
_GEMINI_FREE_DISABLED_LOGGED = False


def gemini_free_api_key() -> str:
    return os.environ.get("GEMINI_API_KEY_FREE", "")


def gemini_free_model() -> str:
    return os.environ.get("GEMINI_FREE_MODEL", "gemini-3.1-flash-lite")


def gemini_free_first_enabled() -> bool:
    val = os.environ.get("GEMINI_FREE_FIRST", "")
    return val.strip().lower() in {"1", "true", "yes", "on"}


def reset_gemini_post_call_count() -> None:
    global GEMINI_POST_CALLS, GEMINI_POST_CALLS_FREE, GEMINI_POST_CALLS_PAID
    global _GEMINI_FREE_COOLDOWN_UNTIL, _GEMINI_FREE_COOLDOWN_STREAK
    global _GEMINI_FREE_DISABLED_FOR_RUN, _GEMINI_FREE_DISABLED_LOGGED
    GEMINI_POST_CALLS = 0
    GEMINI_POST_CALLS_FREE = 0
    GEMINI_POST_CALLS_PAID = 0
    _GEMINI_FREE_WINDOW.clear()
    _GEMINI_FREE_COOLDOWN_UNTIL = 0.0
    _GEMINI_FREE_COOLDOWN_STREAK = 0
    _GEMINI_FREE_DISABLED_FOR_RUN = False
    _GEMINI_FREE_DISABLED_LOGGED = False


def gemini_post_call_count() -> int:
    return GEMINI_POST_CALLS


def gemini_post_call_counts() -> dict[str, int]:
    return {
        "total": GEMINI_POST_CALLS,
        "free": GEMINI_POST_CALLS_FREE,
        "paid": GEMINI_POST_CALLS_PAID,
    }


def gemini_post_call_summary() -> str:
    c = gemini_post_call_counts()
    return f"total {c['total']} | free {c['free']} | paid {c['paid']}"


def _payload_token_estimate(payload: dict) -> int:
    try:
        raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    except Exception:
        raw = str(payload)
    return max(1, len(raw) // 4)


def _gemini_free_limits() -> tuple[int, int]:
    rpm = int(os.environ.get("GEMINI_FREE_RPM_LIMIT", "12") or "12")
    tpm = int(os.environ.get("GEMINI_FREE_TPM_LIMIT", "200000") or "200000")
    return max(0, rpm), max(0, tpm)


def _gemini_free_limiter_allows(payload: dict, now: float | None = None) -> bool:
    now = time.time() if now is None else now
    rpm_limit, tpm_limit = _gemini_free_limits()
    while _GEMINI_FREE_WINDOW and now - _GEMINI_FREE_WINDOW[0][0] >= 60:
        _GEMINI_FREE_WINDOW.pop(0)
    tokens = _payload_token_estimate(payload)
    if len(_GEMINI_FREE_WINDOW) >= rpm_limit:
        return False
    if sum(t for _, t in _GEMINI_FREE_WINDOW) + tokens > tpm_limit:
        return False
    _GEMINI_FREE_WINDOW.append((now, tokens))
    return True


def _gemini_free_available(payload: dict) -> bool:
    if not gemini_free_first_enabled():
        return False
    if not gemini_free_api_key():
        return False
    if _GEMINI_FREE_DISABLED_FOR_RUN:
        return False
    now = time.time()
    if now < _GEMINI_FREE_COOLDOWN_UNTIL:
        return False
    return _gemini_free_limiter_allows(payload, now)


def _gemini_free_mark_quota() -> None:
    global _GEMINI_FREE_COOLDOWN_UNTIL, _GEMINI_FREE_COOLDOWN_STREAK
    global _GEMINI_FREE_DISABLED_FOR_RUN, _GEMINI_FREE_DISABLED_LOGGED
    _GEMINI_FREE_COOLDOWN_UNTIL = time.time() + 60
    _GEMINI_FREE_COOLDOWN_STREAK += 1
    if _GEMINI_FREE_COOLDOWN_STREAK >= 3:
        _GEMINI_FREE_DISABLED_FOR_RUN = True
        if not _GEMINI_FREE_DISABLED_LOGGED:
            print("      gemini free tier quota appears exhausted; using paid key for rest of run",
                  flush=True)
            _GEMINI_FREE_DISABLED_LOGGED = True


def _gemini_free_mark_success() -> None:
    global _GEMINI_FREE_COOLDOWN_STREAK
    _GEMINI_FREE_COOLDOWN_STREAK = 0


def _gemini_is_quota_response(resp: requests.Response) -> bool:
    if resp.status_code == 429:
        return True
    try:
        txt = resp.text or ""
    except Exception:
        txt = ""
    txt = txt.lower()
    return "quota" in txt or "rate limit" in txt or "resource_exhausted" in txt


def _gemini_post_once(session: requests.Session, model: str, key: str,
                      payload: dict, timeout: int, counter: str) -> requests.Response:
    global GEMINI_POST_CALLS, GEMINI_POST_CALLS_FREE, GEMINI_POST_CALLS_PAID
    url = f"{GEMINI_BASE_URL}/models/{model}:generateContent?key={key}"
    GEMINI_POST_CALLS += 1
    if counter == "free":
        GEMINI_POST_CALLS_FREE += 1
    else:
        GEMINI_POST_CALLS_PAID += 1
    return session.post(url, json=payload, timeout=timeout)


def _redact_gemini_error(exc: BaseException) -> str:
    return re.sub(r"([?&]key=)[^&\s)]+", r"\1<redacted>", str(exc))


def _gemini_post_paid(session: requests.Session, model: str, payload: dict,
                      timeout: int = 90) -> dict:
    key = gemini_api_key()
    if not key:
        raise RuntimeError("missing GEMINI_API_KEY")
    for attempt in range(3):
        try:
            resp = _gemini_post_once(session, model, key, payload, timeout, "paid")
            if resp.status_code == 429 or 500 <= resp.status_code < 600:
                if attempt == 2:
                    resp.raise_for_status()
                time.sleep((2 ** attempt) * 2 + random.uniform(0, 1))
                continue
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            print(f"      gemini attempt {attempt+1} failed: {_redact_gemini_error(e)}",
                  flush=True)
            if attempt == 2:
                raise
            time.sleep((2 ** attempt) * 2 + random.uniform(0, 1))
    return {}


def gemini_post(session: requests.Session, model: str, payload: dict,
                timeout: int = 90) -> dict:
    if _gemini_free_available(payload):
        free_key = gemini_free_api_key()
        free_model = gemini_free_model()
        try:
            resp = _gemini_post_once(
                session, free_model, free_key, payload, timeout, "free")
            if _gemini_is_quota_response(resp):
                _gemini_free_mark_quota()
                return _gemini_post_paid(session, model, payload, timeout)
            resp.raise_for_status()
            _gemini_free_mark_success()
            return resp.json()
        except requests.exceptions.RequestException as e:
            print(f"      gemini free attempt failed; falling back to paid: {_redact_gemini_error(e)}",
                  flush=True)
            return _gemini_post_paid(session, model, payload, timeout)
    return _gemini_post_paid(session, model, payload, timeout)


def tier2_grounded_search(session: requests.Session, model: str,
                          municipio: str, site_hint: str,
                          timeout: int = 15,
                          render_page=_DEFAULT_RENDERER) -> list[Candidate]:
    hint = f"O site oficial e: {site_hint}. " if site_hint else ""
    prompt = (
        f"Voce e um investigador de sites oficiais de prefeituras do {UF_NOME} ({UF_SIGLA}), Brasil. "
        f"{hint}"
        f"Encontre no Google as URLs OFICIAIS e ESTAVEIS da prefeitura de "
        f"{municipio} ({UF_NOME}, {UF_SIGLA}, Brasil) para concursos publicos "
        f"e processos seletivos.\n"
        "REGRAS:\n"
        f"- Busque SEMPRE incluindo '{UF_NOME}' ou '{UF_SIGLA}'.\n"
        "- Prefira dominio oficial (.rs.gov.br ou .atende.net).\n"
        "- Queremos a PAGINA INDICE/LISTAGEM (onde se listam varios editais), "
        "NAO um edital individual nem um PDF.\n"
        "- NAO use licitacoes, pregao, compras, chamamento publico.\n"
        "Liste com a URL completa: site oficial; pagina de concursos; pagina de PSS."
    )
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {"temperature": 1.0, "maxOutputTokens": 2048},
    }
    data = gemini_post(session, model, payload, timeout=90)

    cand = (data.get("candidates") or [{}])[0]
    text_response = "\n".join(
        p.get("text", "") for p in (cand.get("content", {}) or {}).get("parts", [])
        if isinstance(p, dict)
    )
    chunks = (cand.get("groundingMetadata", {}) or {}).get("groundingChunks", []) or []

    candidates = []
    seen: set[str] = set()

    # Host filter: only accept URLs from the official site's domain
    hint_host = ""
    if site_hint:
        hint_host = urlparse(site_hint).netloc.lower().lstrip("www.")

    def _t2_host_ok(h: str) -> bool:
        if not hint_host:
            return True
        h = h.lower().lstrip("www.")
        return h == hint_host or h.endswith("." + hint_host)

    # URLs from grounding metadata (real indexed URLs)
    for ch in chunks:
        uri = (ch.get("web", {}) or {}).get("uri", "") if isinstance(ch, dict) else ""
        if not uri:
            continue
        try:
            real = session.get(uri, allow_redirects=True, timeout=timeout).url
            real = clean_url(real)
        except Exception:
            real = clean_url(uri)
        if real and real not in seen:
            host = urlparse(real).netloc.lower()
            if _t2_host_ok(host) and not any(bad in host for bad in BAD_HOSTS) and not is_pdf_or_file(real):
                seen.add(real)
                candidates.append(Candidate(
                    url=real, source="grounding", source_tier="tier2_grounded",
                ))

    # URLs mentioned in text response
    for raw in re.findall(r"https?://[^\s\]\)\"'<>]+", text_response or ""):
        url = clean_url(raw.rstrip(".,;:"))
        if url and url not in seen:
            host = urlparse(url).netloc.lower()
            if _t2_host_ok(host) and not any(bad in host for bad in BAD_HOSTS) and not is_pdf_or_file(url):
                seen.add(url)
                candidates.append(Candidate(
                    url=url, source="grounding", source_tier="tier2_grounded",
                ))

    print(f"      grounding: {len(chunks)} chunks, {len(candidates)} candidate URLs", flush=True)

    # Fetch each to get content for Tier 3
    hydrated_candidates: list[Candidate] = []
    for c in candidates:
        if is_broad_landing(c.url):
            c.fetchable = False
            hydrated_candidates.append(c)
            continue
        hydrated_candidates.append(hydrate_candidate(
            c, municipio, session=session, timeout=timeout,
            official_url=site_hint, render_page=render_page,
        ))

    return hydrated_candidates


def tier2_find_site_grounded(session: requests.Session, model: str,
                             municipio: str, timeout: int = 15) -> Page | None:
    """Discover the official prefeitura domain via grounded search.

    Fallback for when the free slug guesses (Tier 0) miss because the real
    host is non-obvious (abbreviations like pmpf, shortened names like
    caxias, geo-blocked sites, migrations). No fixed rules: Gemini + Google
    find the domain, and we validate it with the same score bar as Tier 0.
    """
    prompt = (
        f"Qual e o site OFICIAL da Prefeitura Municipal de {municipio} "
        f"({UF_NOME}, {UF_SIGLA}, Brasil)?\n"
        "Responda com a URL da PAGINA INICIAL oficial (dominio .rs.gov.br, "
        ".atende.net ou outro dominio oficial da prefeitura). "
        "Nao responda com redes sociais, wikipedia, noticias nem portais de terceiros."
    )
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {"temperature": 0.0, "maxOutputTokens": 1024},
    }
    try:
        data = gemini_post(session, model, payload, timeout=90)
    except Exception as e:
        print(f"      grounded site discovery error: {e}", flush=True)
        return None

    cand = (data.get("candidates") or [{}])[0]
    text_response = "\n".join(
        p.get("text", "") for p in (cand.get("content", {}) or {}).get("parts", [])
        if isinstance(p, dict)
    )
    chunks = (cand.get("groundingMetadata", {}) or {}).get("groundingChunks", []) or []

    # Collect candidate homepage URLs: grounding chunks first (real indexed
    # URLs), then any URL mentioned in the text answer.
    raw_urls: list[str] = []
    for ch in chunks:
        uri = (ch.get("web", {}) or {}).get("uri", "") if isinstance(ch, dict) else ""
        if uri:
            try:
                raw_urls.append(session.get(uri, allow_redirects=True, timeout=timeout).url)
            except Exception:
                raw_urls.append(uri)
    raw_urls.extend(re.findall(r"https?://[^\s\]\)\"'<>]+", text_response or ""))

    # Reduce to candidate base domains, skipping junk hosts.
    seen: set[str] = set()
    base_urls: list[str] = []
    for raw in raw_urls:
        url = clean_url((raw or "").rstrip(".,;:"))
        host = urlparse(url).netloc.lower()
        if not host or host in seen:
            continue
        if any(bad in host for bad in BAD_HOSTS):
            continue
        seen.add(host)
        base_urls.append(f"{urlparse(url).scheme}://{host}/")

    print(f"      grounded site discovery: {len(base_urls)} domain candidates", flush=True)

    best = None
    best_score = -1
    for url in base_urls:
        page = fetch_page(session, url, timeout)
        if not page.ok:
            continue
        score = score_site_page(page, municipio)
        if score > best_score:
            best_score = score
            best = page
    if best and (best_score >= 5
                 or is_matching_official_municipality_domain(best, municipio)):
        migrated = _check_migration(session, best, timeout)
        return migrated or best
    return None


def tier2_directed_bucket_search(session: requests.Session, model: str,
                                 municipio: str, host: str,
                                 bucket_name: str,
                                 timeout: int = 15,
                                 render_page=_DEFAULT_RENDERER) -> list[Candidate]:
    """Targeted grounding search for a specific missing bucket on a known host."""
    prompt = (
        f"Encontre a pagina INDICE/LISTAGEM de {bucket_name} da Prefeitura de "
        f"{municipio} ({UF_NOME}, {UF_SIGLA}, Brasil).\n"
        f"Busque: {bucket_name} site:{host}\n"
        "Queremos a pagina que LISTA VARIOS editais/processos, "
        "NAO um edital individual nem PDF.\n"
        "Liste as URLs encontradas."
    )
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {"temperature": 0.0, "maxOutputTokens": 1024},
    }
    try:
        data = gemini_post(session, model, payload, timeout=90)
    except Exception as e:
        print(f"      directed search error: {e}", flush=True)
        return []

    cand = (data.get("candidates") or [{}])[0]
    text_response = "\n".join(
        p.get("text", "") for p in (cand.get("content", {}) or {}).get("parts", [])
        if isinstance(p, dict)
    )
    chunks = (cand.get("groundingMetadata", {}) or {}).get("groundingChunks", []) or []

    candidates = []
    seen: set[str] = set()
    host_base = host.lower().lstrip("www.")

    def _host_ok(h: str) -> bool:
        h = h.lower().lstrip("www.")
        return h == host_base or h.endswith("." + host_base)

    for ch in chunks:
        uri = (ch.get("web", {}) or {}).get("uri", "") if isinstance(ch, dict) else ""
        if not uri:
            continue
        try:
            real = session.get(uri, allow_redirects=True, timeout=timeout).url
            real = clean_url(real)
        except Exception:
            real = clean_url(uri)
        if real and real not in seen:
            h = urlparse(real).netloc.lower()
            if _host_ok(h) and not any(bad in h for bad in BAD_HOSTS) and not is_pdf_or_file(real):
                seen.add(real)
                candidates.append(Candidate(
                    url=real, source="directed_grounding",
                    source_tier="tier2_directed",
                    bucket_hint=("concursos" if "concurso" in norm(bucket_name)
                                 else "processos"),
                ))

    for raw in re.findall(r"https?://[^\s\]\)\"'<>]+", text_response or ""):
        url = clean_url(raw.rstrip(".,;:"))
        if url and url not in seen:
            h = urlparse(url).netloc.lower()
            if _host_ok(h) and not any(bad in h for bad in BAD_HOSTS) and not is_pdf_or_file(url):
                seen.add(url)
                candidates.append(Candidate(
                    url=url, source="directed_grounding",
                    source_tier="tier2_directed",
                    bucket_hint=("concursos" if "concurso" in norm(bucket_name)
                                 else "processos"),
                ))

    print(f"      directed: {len(candidates)} candidates for {bucket_name}", flush=True)

    hydrated_candidates: list[Candidate] = []
    for c in candidates:
        if is_broad_landing(c.url):
            c.fetchable = False
            hydrated_candidates.append(c)
            continue
        hydrated_candidates.append(hydrate_candidate(
            c, municipio, session=session, timeout=timeout,
            official_url=f"https://{host}/", render_page=render_page,
        ))

    return hydrated_candidates


# ---------------------------------------------------------------------------
# TIER 3: Gemini verifier/selector (discrete decisions, no scores)
# ---------------------------------------------------------------------------
TIER3_DECISIONS = [
    "indice_oficial",
    "indice_oficial_combinado",
    "portal_externo_oficial",
    "detalle_individual_rechazado",
    "licitacao_rechazada",
    "concurso_cultural_rechazado",
    "nao_encontrado",
    "revisar",
]

CONTENT_PAGE_ROLES = set(verdict.PAGE_ROLES)
CONTENT_TIPOS = {"concurso", "pss", "mixto", "incierto"}


def _empty_tier3_result() -> dict:
    return {
        "url_concursos": "",
        "url_processos_seletivos": "",
        "decision_concursos": "nao_encontrado",
        "decision_processos": "nao_encontrado",
        "classificacoes": [],
        "classification_complete": False,
        "razao": "",
    }


def _normalized_candidate_url(url: str) -> str:
    """Canonical key for candidate collision handling, not a quality signal."""
    return _candidate_url_key(url)


def _route_classified_candidates(
        candidates: list[Candidate | CandidateRecord],
        classifications: list[dict],
        selector_picks: dict[str, int | str | None] | None = None) -> dict:
    """Compatibility router that ignores historical AI classifications.

    Candidate dimensions are already immutable. Only selector picks are read,
    and an integer pick is translated to the corresponding existing record ID.
    """
    records = [
        record for candidate in candidates
        if (record := _record_from_candidate(candidate)) is not None
    ]
    result = _empty_tier3_result()
    result["classification_complete"] = len(records) == len(candidates)
    result["selected_resources"] = {}
    reasons = ["classificacoes legacy ignoradas; records pre-adjudicados"]
    selector_picks = selector_picks or {}
    for short_bucket, canonical, url_key, decision_key in (
        ("concursos", "concurso_publico", "url_concursos", "decision_concursos"),
        ("processos", "processo_seletivo", "url_processos_seletivos", "decision_processos"),
    ):
        raw_pick = selector_picks.get(short_bucket)
        picked_id = None
        if isinstance(raw_pick, int) and 0 <= raw_pick < len(candidates):
            picked_record = _record_from_candidate(candidates[raw_pick])
            picked_id = picked_record.candidate_id if picked_record else None
        elif isinstance(raw_pick, str):
            picked_id = raw_pick
        resolved = resolve_selector_pick(records, canonical, picked_id)
        if isinstance(resolved, SelectedResource):
            result["selected_resources"][short_bucket] = resolved
            result[url_key] = resolved.candidate.final_url
            result[decision_key] = resolved.candidate.decision
            reasons.append(f"{short_bucket}: {resolved.reason}")
        else:
            result[decision_key] = resolved.decision
            reasons.append(f"{short_bucket}: {resolved.reason}")
    result["razao"] = "; ".join(reasons)
    return result


def _canonical_bucket(bucket: str) -> str:
    return {
        "concursos": "concurso_publico",
        "concurso_publico": "concurso_publico",
        "processos": "processo_seletivo",
        "processo_seletivo": "processo_seletivo",
        "combinado": "combinado",
    }.get(bucket, "")


def _record_from_candidate(
        candidate: Candidate | CandidateRecord,
        ) -> CandidateRecord | None:
    if isinstance(candidate, CandidateRecord):
        return candidate
    return candidate.record


def _record_supports_bucket(record: CandidateRecord, bucket: str) -> bool:
    canonical = _canonical_bucket(bucket)
    return bool(canonical and (
        record.bucket == canonical or record.bucket == "combinado"
    ))


def _eligible_records_for_bucket(
        records: list[CandidateRecord], bucket: str,
        ) -> list[CandidateRecord]:
    canonical = _canonical_bucket(bucket)
    if canonical not in {"concurso_publico", "processo_seletivo"}:
        return []
    eligible = [
        record for record in records
        if record.eligible and _record_supports_bucket(record, canonical)
    ]
    # Contract preference: a combined index is eligible only when no dedicated
    # surface exists for this bucket. This is discrete precedence, not scoring.
    specific = [record for record in eligible if record.bucket == canonical]
    return specific or eligible


def _review_final(bucket: str, reason: str, *, decision: str = "revisar",
                  candidate_id: str = "",
                  record: CandidateRecord | None = None) -> FinalDecision:
    final = FinalDecision(
        bucket=_canonical_bucket(bucket) or bucket,
        status="revisar",
        decision=decision,
        url="",
        candidate_id=candidate_id,
        reason=reason or "revisar sin razon recibida",
    )
    LOGGER.info("final_decision %s", json.dumps({
        "bucket": final.bucket,
        "status": final.status,
        "decision": final.decision,
        "url": final.url,
        "candidate_id": final.candidate_id,
        "authority": record.authority if record else "sin_record",
        "identity": record.identity if record else "sin_record",
        "page_role": record.page_role if record else "sin_record",
        "evidence_state": record.evidence_state if record else "sin_record",
        "record_bucket": record.bucket if record else "sin_record",
        "reason": final.reason,
    }, ensure_ascii=False, sort_keys=True))
    return final


def resolve_selector_pick(
        records: list[CandidateRecord], bucket: str,
        candidate_id: str | None,
        ) -> SelectedResource | FinalDecision:
    """Validate a Tier3 ID against already adjudicated, eligible records."""
    canonical = _canonical_bucket(bucket)
    if canonical not in {"concurso_publico", "processo_seletivo"}:
        return _review_final(bucket, f"bucket desconocido: {bucket}")
    pool = _eligible_records_for_bucket(records, canonical)
    if not pool:
        return _review_final(canonical, "sin CandidateRecord elegible para el bucket")

    unique: dict[str, CandidateRecord] = {}
    for record in pool:
        unique.setdefault(record.candidate_id, record)

    if candidate_id:
        by_all_id = {record.candidate_id: record for record in records}
        if candidate_id not in by_all_id:
            return _review_final(
                canonical, f"Tier3 devolvio candidate_id inexistente: {candidate_id}",
                candidate_id=candidate_id,
            )
        if candidate_id not in unique:
            return _review_final(
                canonical,
                f"Tier3 devolvio candidate_id no elegible o incompatible: {candidate_id}",
                candidate_id=candidate_id,
            )
        return SelectedResource(canonical, unique[candidate_id])

    if len(unique) == 1:
        selected = next(iter(unique.values()))
        reason = (
            "duplicado elegible colapsado deterministicamente por candidate_id"
            if len(pool) > 1 else
            "unico CandidateRecord elegible; seleccion determinista"
        )
        return SelectedResource(canonical, selected, reason)
    return _review_final(
        canonical,
        "empate entre CandidateRecord elegibles: Tier3 no devolvio candidate_id",
    )


def derive_final_decision(selected: SelectedResource) -> FinalDecision:
    """Derive the final state without fetching or re-adjudicating evidence."""
    record = selected.candidate
    canonical = _canonical_bucket(selected.bucket)
    provenance_reason = (
        f"candidate_id={record.candidate_id}; tier={record.tier}; "
        f"requested_url={record.requested_url}; final_url={record.final_url}"
    )
    if canonical not in {"concurso_publico", "processo_seletivo"}:
        return _review_final(selected.bucket, f"bucket desconocido: {selected.bucket}")
    accepted = {
        "indice_oficial", "indice_oficial_combinado", "portal_externo_oficial",
    }
    blockers = []
    if record.authority != "confirmada":
        blockers.append(f"authority={record.authority}")
    if record.identity != "confirmada":
        blockers.append(f"identity={record.identity}")
    if not record.accessible or record.evidence_state not in {"completa", "renderizada"}:
        blockers.append(f"evidence_state={record.evidence_state}")
    if record.decision not in accepted:
        blockers.append(f"decision={record.decision}")
    if not _record_supports_bucket(record, canonical):
        blockers.append(f"bucket_incompatible={record.bucket}")
    if blockers:
        return _review_final(
            canonical,
            f"{record.reason}; {provenance_reason}; " + "; ".join(blockers),
            decision=record.decision,
            candidate_id=record.candidate_id,
            record=record,
        )

    reason = (
        f"{record.reason}; {selected.reason}; {provenance_reason}; "
        "snapshot preservado sin refetch"
    )
    final = FinalDecision(
        bucket=canonical,
        status="confirmado",
        decision=record.decision,
        url=record.final_url,
        candidate_id=record.candidate_id,
        reason=reason,
    )
    LOGGER.info("final_decision %s", json.dumps({
        "bucket": final.bucket,
        "status": final.status,
        "decision": final.decision,
        "url": final.url,
        "candidate_id": final.candidate_id,
        "authority": record.authority,
        "identity": record.identity,
        "page_role": record.page_role,
        "evidence_state": record.evidence_state,
        "record_bucket": record.bucket,
        "reason": final.reason,
    }, ensure_ascii=False, sort_keys=True))
    return final


# Public Tier 3 seam: selector-only contract.
def tier3_classify_and_pick(session: requests.Session, model: str,
                            municipio: str,
                            candidates: list[Candidate | CandidateRecord],
                            timeout: int = 30) -> dict:
    """Tier 3 selector: return IDs only; never classify or mutate records."""
    records = [
        record for candidate in candidates
        if (record := _record_from_candidate(candidate)) is not None
    ]
    result = _empty_tier3_result()
    result["classification_complete"] = True
    result["selected_resources"] = {}
    result["classificacoes"] = [{
        "candidate_id": record.candidate_id,
        "url": record.final_url,
        "authority": record.authority,
        "identity": record.identity,
        "page_role": record.page_role,
        "evidence_state": record.evidence_state,
        "bucket": record.bucket,
        "decision_code": record.decision,
        "razao": record.reason,
    } for record in records]
    if not records:
        result["razao"] = "Tier3 sin CandidateRecord adjudicado"
        return result

    pools = {
        "concursos": _eligible_records_for_bucket(records, "concurso_publico"),
        "processos": _eligible_records_for_bucket(records, "processo_seletivo"),
    }
    ai_picks: dict[str, str | None] = {"concursos": None, "processos": None}
    needs_ai = {
        bucket: pool for bucket, pool in pools.items()
        if len({record.candidate_id for record in pool}) > 1
    }
    ai_reason = ""
    if needs_ai:
        items = []
        seen_ids = set()
        for pool in needs_ai.values():
            for record in pool:
                if record.candidate_id in seen_ids:
                    continue
                seen_ids.add(record.candidate_id)
                items.append({
                    "candidate_id": record.candidate_id,
                    "url": record.final_url,
                    "bucket": record.bucket,
                    "decision": record.decision,
                    "title": record.evidence_snapshot.title[:160],
                    "content_preview": record.evidence_snapshot.text[:2400],
                })
        prompt = (
            f"Prefeitura de {municipio} ({UF_NOME}, {UF_SIGLA}). "
            "As candidatas abaixo JA foram adjudicadas deterministicamente e "
            "todas as dimensoes sao imutaveis. Atue SOMENTE como seletor. "
            "Escolha o candidate_id do indice mais amplo, estavel e canonico "
            "para cada bucket. Nao classifique, confirme, altere dimensoes nem "
            "invente IDs. Paginas especificas ja tiveram precedencia sobre "
            "paginas combinadas no pool.\n\n"
            f"Candidatas: {json.dumps(items, ensure_ascii=False)}\n\n"
            "Responda JSON: {\"candidate_id_concursos\": \"v1:...\" ou null, "
            "\"candidate_id_processos\": \"v1:...\" ou null, "
            "\"razao\": \"curta\"}."
        )
        payload = {
            "contents": [{"role": "user", "parts": [{"text": prompt}]}],
            "generationConfig": {
                "temperature": 0.0, "maxOutputTokens": 1024,
                "responseMimeType": "application/json",
            },
        }
        try:
            data = gemini_post(session, model, payload, timeout=60)
            text = "\n".join(
                part.get("text", "")
                for part in data["candidates"][0]["content"]["parts"]
            )
            raw = json.loads(text)
            ai_picks = {
                "concursos": raw.get("candidate_id_concursos"),
                "processos": raw.get("candidate_id_processos"),
            }
            ai_reason = str(raw.get("razao", "")).strip()
        except Exception as exc:
            ai_reason = f"Tier3 selector no pudo elegir: {type(exc).__name__}"

    reasons = []
    for short_bucket, canonical, url_key, decision_key in (
        ("concursos", "concurso_publico", "url_concursos", "decision_concursos"),
        ("processos", "processo_seletivo", "url_processos_seletivos", "decision_processos"),
    ):
        pick = ai_picks[short_bucket] if short_bucket in needs_ai else None
        resolved = resolve_selector_pick(records, canonical, pick)
        if isinstance(resolved, SelectedResource):
            result["selected_resources"][short_bucket] = resolved
            result[url_key] = resolved.candidate.final_url
            result[decision_key] = resolved.candidate.decision
            reasons.append(f"{short_bucket}: {resolved.reason}")
        else:
            result[decision_key] = resolved.decision
            reasons.append(f"{short_bucket}: {resolved.reason}")
    if ai_reason:
        reasons.append(f"selector: {ai_reason}")
    result["razao"] = "; ".join(reasons)
    return result


# ---------------------------------------------------------------------------
# TIER 4: Playwright directed navigation
# ---------------------------------------------------------------------------
_BROWSER = None


def _find_chromium() -> str | None:
    import glob
    for pat in ("/opt/pw-browsers/chromium-*/chrome-linux/chrome",
                "/opt/pw-browsers/chromium-*/chrome-linux/headless_shell"):
        hits = sorted(glob.glob(pat))
        if hits:
            return hits[-1]
    return None


def _get_browser():
    global _BROWSER
    if _BROWSER is None:
        from playwright.sync_api import sync_playwright
        pw = sync_playwright().start()
        chrome_path = _find_chromium()
        # Hardening for proxied/headless environments. Secure DNS (DoH) probes
        # to dns.google are unreachable behind an egress proxy and flood the
        # net stack, so turn DoH and background networking off and let the
        # proxy resolve names via CONNECT.
        launch_args = [
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--dns-over-https-mode=off",
            "--disable-features=DnsOverHttps,DnsOverHttpsUpgrade",
            "--disable-background-networking",
            "--disable-component-update",
            "--no-pings",
            "--disable-blink-features=AutomationControlled",
            "--lang=pt-BR",
        ]
        proxy_url = os.environ.get("HTTPS_PROXY") or os.environ.get("https_proxy")
        if proxy_url:
            launch_args.append(f"--proxy-server={proxy_url}")
            # A re-terminating egress proxy forges a per-host cert chain and
            # cannot complete BoringSSL's TLS 1.3 ClientHello (post-quantum
            # keyshare). Accept the forged cert and cap at TLS 1.2 so the
            # tunnel handshake succeeds. Only applied when a proxy is present;
            # direct-egress (production) keeps modern TLS.
            launch_args += [
                "--ignore-certificate-errors",
                "--test-type",
                "--ssl-version-max=tls1.2",
            ]
        _BROWSER = pw.chromium.launch(
            headless=True,
            executable_path=chrome_path,
            args=launch_args,
        )
    return _BROWSER


def render_page_sync(url: str) -> RenderedPage | None:
    """Render one URL with the shared sync browser and return the final DOM."""
    try:
        browser = _get_browser()
    except Exception as e:
        print(f"      challenge render unavailable: {e}", flush=True)
        return None
    context = None
    try:
        context = new_browser_context(browser, ignore_https_errors=True)
        browser_page = context.new_page()
        response = browser_page.goto(
            url, wait_until="domcontentloaded", timeout=20000,
        )
        browser_page.wait_for_timeout(2000)
        links = browser_page.eval_on_selector_all(
            "a[href]",
            "els => els.map(el => [el.href, (el.innerText || '').trim()])",
        )
        return RenderedPage(
            html=browser_page.content(),
            text=browser_page.locator("body").inner_text(),
            title=browser_page.title(),
            requested_url=url,
            final_url=browser_page.url,
            status=response.status if response is not None else None,
            links=tuple((href, text) for href, text in links),
        )
    except Exception as e:
        print(f"      challenge render error: {e}", flush=True)
        return None
    finally:
        if context is not None:
            try:
                context.close()
            except Exception:
                pass


def _render_page_links(url: str, timeout: int = 20) -> list[tuple[str, str]]:
    """Load a JS-rendered page in a headless browser and return its <a> links.

    Used only for SPA shells in Tier 1, where the served HTML has no usable menu.
    Reuses the shared browser; returns [] if the browser is unavailable so the
    caller silently keeps the (empty) static result.
    """
    try:
        browser = _get_browser()
    except Exception as e:
        print(f"      SPA render unavailable: {e}", flush=True)
        return []
    context = None
    try:
        context = new_browser_context(
            browser,
            ignore_https_errors=True,
        )
        page = context.new_page()
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(2000)
        links = page.evaluate("""() => {
            const results = [];
            document.querySelectorAll('a[href]').forEach(el => {
                results.push({href: el.href, text: (el.innerText || '').trim()});
            });
            return results;
        }""")
        out: list[tuple[str, str]] = []
        seen: set[str] = set()
        for link in links:
            href = link.get("href", "")
            if href.startswith("http") and href not in seen:
                seen.add(href)
                out.append((href, link.get("text", "")))
        return out
    except Exception as e:
        print(f"      SPA render error: {e}", flush=True)
        return []
    finally:
        if context is not None:
            try:
                context.close()
            except Exception:
                pass


def _tier4_candidates_from_links(
        relevant_links: list[tuple[str, str]], municipio: str,
        render_page) -> list[Candidate]:
    """Render and validate every Tier 4 link independently, without refetching."""
    candidates: list[Candidate] = []
    for href, menu_text in relevant_links:
        if is_broad_landing(href) or is_pdf_or_file(href):
            continue
        try:
            snapshot = render_page(href)
        except Exception as e:
            print(f"      playwright candidate render error: {e}", flush=True)
            continue
        if not snapshot:
            continue
        candidates.append(hydrate_candidate(
            Candidate(
                url=href, source="playwright", menu_text=menu_text,
                source_tier="tier4",
            ),
            municipio, evidence=snapshot,
        ))
    return candidates


def tier4_playwright_collect(url: str, municipio: str) -> list[Candidate]:
    """Navigate the site like a human: open menus, follow relevant links."""
    try:
        browser = _get_browser()
    except Exception as e:
        print(f"      playwright unavailable: {e}", flush=True)
        return []

    candidates = []
    all_keywords = []
    for kws in BUCKET_KEYWORDS.values():
        all_keywords.extend(kws)

    context = None
    try:
        context = new_browser_context(
            browser,
            ignore_https_errors=True,
        )
        page = context.new_page()
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(2000)

        # Extract all links from the rendered page
        links = page.evaluate("""() => {
            const results = [];
            document.querySelectorAll('a[href]').forEach(el => {
                results.push({href: el.href, text: (el.innerText || '').trim()});
            });
            return results;
        }""")

        seen = set()
        relevant_links = []
        container_links = []

        for link in links:
            href = link.get("href", "")
            text = link.get("text", "")
            if not href.startswith("http") or href in seen:
                continue
            host = urlparse(href).netloc.lower()
            if any(bad in host for bad in BAD_HOSTS):
                continue
            text_n = norm(text)
            href_n = norm(unquote(urlparse(href).path))

            if any(kw in text_n or kw in href_n for kw in all_keywords):
                seen.add(href)
                relevant_links.append((href, text))
            elif any(kw in text_n or kw in href_n for kw in CONTAINER_KEYWORDS):
                container_links.append((href, text))

        # Follow container links one level deep
        for container_href, container_text in container_links[:5]:
            if container_href in seen:
                continue
            try:
                page2 = context.new_page()
                page2.goto(container_href, wait_until="domcontentloaded", timeout=15000)
                page2.wait_for_timeout(1500)
                sub_links = page2.evaluate("""() => {
                    const results = [];
                    document.querySelectorAll('a[href]').forEach(el => {
                        results.push({href: el.href, text: (el.innerText || '').trim()});
                    });
                    return results;
                }""")
                for sl in sub_links:
                    sh = sl.get("href", "")
                    st = sl.get("text", "")
                    if not sh.startswith("http") or sh in seen:
                        continue
                    st_n = norm(st)
                    sh_n = norm(unquote(urlparse(sh).path))
                    if any(kw in st_n or kw in sh_n for kw in all_keywords):
                        seen.add(sh)
                        relevant_links.append((sh, f"{container_text} > {st}"))
                page2.close()
            except Exception:
                pass

        def render_in_context(href: str) -> EvidenceSnapshot | None:
            candidate_page = context.new_page()
            try:
                response = candidate_page.goto(
                    href, wait_until="domcontentloaded", timeout=15000,
                )
                candidate_page.wait_for_timeout(1500)
                candidate_links = candidate_page.eval_on_selector_all(
                    "a[href]",
                    "els => els.map(el => [el.href, (el.innerText || '').trim()])",
                )
                return EvidenceSnapshot(
                    html=candidate_page.content(),
                    text=candidate_page.locator("body").inner_text(),
                    title=candidate_page.title(),
                    requested_url=href,
                    final_url=candidate_page.url,
                    status=response.status if response is not None else None,
                    source="playwright",
                    links=tuple(
                        (link_href, link_text)
                        for link_href, link_text in candidate_links
                    ),
                )
            finally:
                candidate_page.close()

        candidates = _tier4_candidates_from_links(
            relevant_links, municipio, render_in_context,
        )

        page.close()
    except Exception as e:
        print(f"      playwright error: {e}", flush=True)
    finally:
        if context is not None:
            try:
                context.close()
            except Exception:
                pass

    print(f"      playwright: {len(candidates)} candidates found", flush=True)

    return candidates


# ---------------------------------------------------------------------------
# Main pipeline: process one municipality
# ---------------------------------------------------------------------------
# Compatibility export consumed by the read-only Fase 2 auditor. It identifies
# textual listing events; it is not used to score or finalize CandidateRecords.
LISTING_SIGNALS = [
    r"\b\d{1,3}/20[12]\d\b",
    r"edital\s+n",
    r"inscri[cç][oõ]es\s+(aberta|encerrada)",
    r"resultado\s+(final|parcial|preliminar)",
    r"homologa[cç][aã]o",
    r"retifica[cç][aã]o",
]
LISTING_RE = re.compile("|".join(LISTING_SIGNALS), re.I)


def _deterministic_verify(url: str, bucket: str,
                          all_candidates: list[Candidate | CandidateRecord]) -> bool:
    """Compatibility query over the existing immutable adjudication."""
    record = next((
        record for candidate in all_candidates
        if (record := _record_from_candidate(candidate)) is not None
        and _normalized_candidate_url(record.final_url)
        == _normalized_candidate_url(url)
    ), None)
    return bool(
        record and record.eligible and _record_supports_bucket(record, bucket)
    )


def batch_gemini_verify(*_args, **_kwargs):
    """Removed compatibility symbol; batch has no independent adjudicator."""
    raise RuntimeError(
        "batch_gemini_verify eliminado: use build_candidate_record + "
        "derive_final_decision"
    )


def _batch_verify_uncertain_results(
        session: requests.Session, model: str, results: list[MunicipioResult],
        *, timeout: int = 15, use_playwright: bool = True,
        ) -> tuple[list[dict], dict[str, MunicipioResult], dict[str, tuple[str, str]]]:
    """Apply the common closure to selected or legacy probable results.

    Selected resources are finalized from their exact immutable record. Legacy
    URL-only rows may perform one compatibility fetch to obtain evidence, but
    that evidence is immediately converted by ``build_candidate_record`` and
    finalized by the same ``derive_final_decision`` path. There is no batch AI
    verdict and no refetch of a selected CandidateRecord.
    """
    to_verify: list[dict] = []
    verify_index: dict[str, MunicipioResult] = {}
    for result in results:
        for bucket, url, confidence in (
            ("concursos", result.url_concursos, result.confianza_concursos),
            ("processos", result.url_processos_seletivos,
             result.confianza_processos),
        ):
            if confidence != "probable" or not url:
                continue
            item = {
                "municipio": result.municipio, "bucket": bucket,
                "url": url, "title": "", "preview": "",
                "site_base": result.site_base,
            }
            to_verify.append(item)
            verify_index[f"{result.municipio}|{bucket}"] = result

    verdicts: dict[str, tuple[str, str]] = {}
    for item in to_verify:
        key = f"{item['municipio']}|{item['bucket']}"
        result = verify_index[key]
        canonical = _canonical_bucket(item["bucket"])
        selected = result.selected_resources.get(item["bucket"])
        if selected is None:
            association = result.selected_evidence.get(item["bucket"])
            if association is not None:
                association_record = _record_from_candidate(association.candidate)
                if association_record is None:
                    legacy_candidate = association.candidate
                    association_record = build_candidate_record(
                        requested_url=association.snapshot.requested_url or item["url"],
                        source=getattr(legacy_candidate, "source", "legacy_snapshot"),
                        tier=getattr(legacy_candidate, "source_tier", "legacy"),
                        municipio=item["municipio"],
                        bucket_hint=getattr(
                            legacy_candidate, "bucket_hint", item["bucket"],
                        ),
                        evidence=association.snapshot,
                        menu_text=getattr(legacy_candidate, "menu_text", ""),
                        provenance=getattr(legacy_candidate, "provenance", ()),
                    )
                if association_record is not None:
                    selected = SelectedResource(
                        canonical, association_record,
                        "adaptador selected_evidence con instancia exacta",
                    )
        if selected is None:
            page = fetch_page(session, item["url"], timeout=timeout)
            snapshot = _snapshot_from_page(item["url"], page)
            legacy_record = build_candidate_record(
                requested_url=item["url"],
                source="legacy",
                tier="legacy",
                municipio=item["municipio"],
                bucket_hint=item["bucket"],
                evidence=snapshot,
            )
            selected = SelectedResource(
                canonical, legacy_record,
                "adaptador legacy: evidencia obtenida una vez y adjudicada centralmente",
            )
            result.selected_resources[item["bucket"]] = selected
            result.selected_evidence[item["bucket"]] = BucketCandidateEvidence(
                item["bucket"], legacy_record, snapshot,
            )

        final = derive_final_decision(selected)
        result.final_decisions[item["bucket"]] = final
        item["title"] = selected.candidate.evidence_snapshot.title[:150]
        item["preview"] = selected.candidate.evidence_snapshot.text[:400]
        verdicts[key] = (final.status, final.reason)
        if item["bucket"] == "concursos":
            result.confianza_concursos = final.status
            result.url_concursos = final.url
            result.notes = (result.notes + f"; verify_c: {final.reason}").strip("; ")
        else:
            result.confianza_processos = final.status
            result.url_processos_seletivos = final.url
            result.notes = (result.notes + f"; verify_p: {final.reason}").strip("; ")

    return to_verify, verify_index, verdicts


def grounded_verify_one(*_args, **_kwargs):
    """Removed compatibility symbol; grounding discovers but never adjudicates."""
    raise RuntimeError(
        "grounded_verify_one eliminado como autoridad de decision; "
        "use build_candidate_record + derive_final_decision"
    )


@dataclass
class MunicipioResult:
    municipio: str
    site_base: str = ""
    url_concursos: str = ""
    url_processos_seletivos: str = ""
    method: str = ""
    notes: str = ""
    tier_concursos: str = ""
    tier_processos: str = ""
    razao: str = ""
    confianza_concursos: str = ""
    confianza_processos: str = ""
    urls_extras_concursos: str = ""
    urls_extras_processos: str = ""
    selected_evidence: dict[str, BucketCandidateEvidence] = field(
        default_factory=dict, repr=False,
    )
    selected_resources: dict[str, SelectedResource] = field(
        default_factory=dict, repr=False,
    )
    final_decisions: dict[str, FinalDecision] = field(
        default_factory=dict, repr=False,
    )


def process_municipio(session: requests.Session, municipio: str,
                      model: str, timeout: int = 15,
                      use_playwright: bool = True,
                      render_page=_DEFAULT_RENDERER) -> MunicipioResult:
    result = MunicipioResult(municipio=municipio)
    tiers_used = []
    all_candidates: list[Candidate] = []
    challenge_renderer = render_page
    if challenge_renderer is _DEFAULT_RENDERER:
        challenge_renderer = render_page_sync if use_playwright else None

    # --- TIER 0: Find site base (free slug guesses) ---
    print(f"  [{municipio}] Tier 0: finding site...", flush=True)
    home = tier0_find_site(
        session, municipio, timeout, render_page=challenge_renderer,
    )
    if home:
        tiers_used.append("t0")
    elif gemini_api_key():
        # Free path missed (non-obvious domain, geo-block, migration):
        # let grounded search discover the official domain.
        print(f"    Tier 0 free miss; grounded site discovery...", flush=True)
        home = tier2_find_site_grounded(session, model, municipio, timeout)
        if home:
            tiers_used.append("t2site")

    if not home:
        result.notes = "site_not_found"
        result.method = "+".join(tiers_used) + ("+" if tiers_used else "") + "tier0_failed"
        return result
    result.site_base = clean_url(home.url)
    print(f"    site: {result.site_base}", flush=True)

    # --- TIER 1: Free link discovery ---
    print(f"    Tier 1: scanning links...", flush=True)
    # SPA shell: the served HTML has no usable menu (rendered client-side).
    # Render it once with the browser so Tier 1 can see the real links. Gated on
    # use_playwright; normal sites are untouched and stay cheap.
    if home.is_spa and use_playwright and len(home.links) < 8:
        print(f"    Tier 1: SPA shell detected, rendering menu with browser...", flush=True)
        rendered = _render_page_links(home.url, timeout)
        if rendered:
            existing = {h for h, _ in home.links}
            home.links.extend((h, t) for h, t in rendered if h not in existing)
            print(f"    Tier 1: rendered {len(rendered)} links from SPA menu", flush=True)
            tiers_used.append("t1spa")
    t1_candidates = tier1_collect_candidates(
        session, home, municipio, timeout, render_page=challenge_renderer,
    )
    all_candidates.extend(t1_candidates)
    fetchable_t1 = [c for c in t1_candidates if c.fetchable]
    print(f"    Tier 1: {len(fetchable_t1)} fetchable candidates from {len(t1_candidates)} found", flush=True)
    tiers_used.append("t1")

    # --- TIER 3 on Tier 1 candidates (if we have any) ---
    chosen = {"url_concursos": "", "url_processos_seletivos": ""}
    bucket_tier = {"url_concursos": "", "url_processos_seletivos": ""}
    bucket_decision = {
        "url_concursos": "nao_encontrado",
        "url_processos_seletivos": "nao_encontrado",
    }
    bucket_selected: dict[str, SelectedResource] = {}
    razones: list[str] = []

    def _record(picked: dict, tier_label: str) -> None:
        """Record a fresh selection over the complete accumulated candidate set."""
        # A malformed/truncated response is not a new content verdict. Preserve
        # the last complete selection instead of erasing a correct earlier tier.
        if not picked.get("classification_complete"):
            razones.append(f"[{tier_label}] tier3 incomplete; previous selection preserved")
            return
        for key, decision_key in (
            ("url_concursos", "decision_concursos"),
            ("url_processos_seletivos", "decision_processos"),
        ):
            previous = chosen[key]
            selected = picked.get(key, "")
            chosen[key] = selected
            bucket_decision[key] = picked.get(decision_key, "nao_encontrado")
            short_bucket = "concursos" if key == "url_concursos" else "processos"
            picked_resource = (picked.get("selected_resources") or {}).get(short_bucket)
            if isinstance(picked_resource, SelectedResource):
                bucket_selected[short_bucket] = picked_resource
            elif not selected:
                bucket_selected.pop(short_bucket, None)
            if selected and selected != previous:
                bucket_tier[key] = tier_label
            elif not selected:
                bucket_tier[key] = ""
        if picked.get("razao"):
            razones.append(f"[{tier_label}] {picked['razao']}")

    if fetchable_t1 and gemini_api_key():
        print(f"    Tier 3: selecting among {len(fetchable_t1)} Tier 1 records...", flush=True)
        picked = tier3_classify_and_pick(session, model, municipio, fetchable_t1, timeout)
        _record(picked, "t1")
        tiers_used.append("t3")

    # --- TIER 1.5: probe known CMS index paths (FALLBACK only) ---
    # Fires only for buckets Tier 1 + Tier 3 left empty, so it never competes
    # with an index already discovered via the menu (which the golden treats as
    # canonical — e.g. an external delegated portal). The content guard inside
    # the probe keeps soft-404s / generic CMS fallbacks out of Tier 3.
    if (not chosen["url_concursos"] or not chosen["url_processos_seletivos"]) \
            and gemini_api_key():
        existing_urls = {c.url for c in all_candidates}
        probe_cands = _probe_known_index_paths(
            session, home, existing_urls, municipio, timeout,
        )
        if probe_cands:
            all_candidates.extend(probe_cands)
            tiers_used.append("probe")
            print(f"    Probe: {len(probe_cands)} known-path candidate(s)...", flush=True)
            picked = tier3_classify_and_pick(
                session, model, municipio,
                [c for c in all_candidates if c.fetchable], timeout,
            )
            _record(picked, "probe")

    # --- TIER 2: Grounded search (only for missing buckets) ---
    missing_buckets = []
    if not chosen["url_concursos"]:
        missing_buckets.append("concursos publicos")
    if not chosen["url_processos_seletivos"]:
        missing_buckets.append("processos seletivos")

    if missing_buckets and gemini_api_key():
        print(f"    Tier 2: grounded search (missing: {', '.join(missing_buckets)})...", flush=True)
        try:
            t2_candidates = tier2_grounded_search(
                session, model, municipio, result.site_base, timeout,
                render_page=challenge_renderer,
            )
            # Filter out candidates we already have
            existing_urls = {c.url for c in all_candidates}
            new_t2 = [c for c in t2_candidates if c.url not in existing_urls]
            all_candidates.extend(new_t2)
            tiers_used.append("t2")

            # Run Tier 3 on new candidates (plus any unfilled from before)
            fetchable_new = [c for c in new_t2 if c.fetchable]
            if fetchable_new:
                print(f"    Tier 3: selecting among {len(fetchable_new)} grounded records...", flush=True)
                picked = tier3_classify_and_pick(
                    session, model, municipio,
                    [c for c in all_candidates if c.fetchable], timeout,
                )
                _record(picked, "t2")
        except Exception as e:
            print(f"    Tier 2 error: {e}", flush=True)
            tiers_used.append("t2_err")

    # --- Directed grounding per bucket ---
    # If general grounding didn't find a bucket, search specifically for it
    # on the known host (e.g. "processos seletivos site:pmpf.rs.gov.br").
    dir_missing = []
    if not chosen["url_concursos"]:
        dir_missing.append(("url_concursos", "concursos publicos"))
    if not chosen["url_processos_seletivos"]:
        dir_missing.append(("url_processos_seletivos", "processos seletivos"))

    if dir_missing and result.site_base and gemini_api_key():
        host = urlparse(result.site_base).netloc
        for bucket_key, bucket_name in dir_missing:
            print(f"    Directed grounding: {bucket_name} on {host}...", flush=True)
            try:
                t2d = tier2_directed_bucket_search(
                    session, model, municipio, host, bucket_name, timeout,
                    render_page=challenge_renderer,
                )
                existing_urls = {c.url for c in all_candidates}
                new_d = [c for c in t2d if c.url not in existing_urls]
                all_candidates.extend(new_d)
                fetchable_d = [c for c in new_d if c.fetchable]
                if fetchable_d:
                    picked = tier3_classify_and_pick(
                        session, model, municipio,
                        [c for c in all_candidates if c.fetchable], timeout,
                    )
                    _record(picked, "t2dir")
                tiers_used.append("t2dir")
            except Exception as e:
                print(f"    Directed grounding error: {e}", flush=True)

    # --- TIER 4: Playwright (last resort for still-missing buckets) ---
    still_missing = []
    if not chosen["url_concursos"]:
        still_missing.append("concursos")
    if not chosen["url_processos_seletivos"]:
        still_missing.append("processos")

    if still_missing and use_playwright:
        print(f"    Tier 4: playwright navigation (missing: {', '.join(still_missing)})...", flush=True)
        try:
            t4_candidates = tier4_playwright_collect(result.site_base, municipio)
            existing_urls = {c.url for c in all_candidates}
            new_t4 = [c for c in t4_candidates if c.url not in existing_urls]
            all_candidates.extend(new_t4)
            tiers_used.append("t4")

            fetchable_t4 = [c for c in new_t4 if c.fetchable]
            if fetchable_t4 and gemini_api_key():
                print(f"    Tier 3: selecting among {len(fetchable_t4)} playwright records...", flush=True)
                picked = tier3_classify_and_pick(
                    session, model, municipio,
                    [c for c in all_candidates if c.fetchable], timeout,
                )
                _record(picked, "t4")
        except Exception as e:
            print(f"    Tier 4 error: {e}", flush=True)

    # --- Assemble result ---
    result.url_concursos = chosen.get("url_concursos", "")
    result.url_processos_seletivos = chosen.get("url_processos_seletivos", "")
    result.method = "+".join(tiers_used)
    result.tier_concursos = bucket_tier["url_concursos"]
    result.tier_processos = bucket_tier["url_processos_seletivos"]
    result.razao = " | ".join(razones)

    # --- One closure: SelectedResource -> FinalDecision (no fetch/re-adjudication) ---
    for short_bucket, canonical, url_attr, confidence_attr, decision_key in (
        ("concursos", "concurso_publico", "url_concursos",
         "confianza_concursos", "url_concursos"),
        ("processos", "processo_seletivo", "url_processos_seletivos",
         "confianza_processos", "url_processos_seletivos"),
    ):
        selected = bucket_selected.get(short_bucket)
        selected_url = getattr(result, url_attr)
        if selected is None and selected_url:
            exact_record = next((
                record for candidate in all_candidates
                if (record := _record_from_candidate(candidate)) is not None
                and _normalized_candidate_url(record.final_url)
                == _normalized_candidate_url(selected_url)
            ), None)
            if exact_record is not None:
                selected = SelectedResource(
                    canonical, exact_record,
                    "seleccion legacy de Tier3 enlazada a la instancia exacta por URL",
                )
        if selected is not None:
            final = derive_final_decision(selected)
            result.selected_resources[short_bucket] = selected
            result.final_decisions[short_bucket] = final
            setattr(result, url_attr, final.url)
            setattr(result, confidence_attr, final.status)
            razones.append(f"[final:{short_bucket}] {final.reason}")
        else:
            decision = bucket_decision[decision_key]
            reason = (
                f"sin recurso seleccionado; decision={decision}; tiers agotados"
            )
            final = _review_final(canonical, reason, decision=decision)
            result.final_decisions[short_bucket] = final
            setattr(result, confidence_attr, (
                "revisar" if decision != "nao_encontrado" else ""
            ))
            razones.append(f"[final:{short_bucket}] {final.reason}")
    result.razao = " | ".join(razones)

    # Downgrade to "revisar" when site not found or all tiers exhausted
    antibot_block = False
    if not result.url_concursos and not result.url_processos_seletivos:
        if any(c.fetchable for c in all_candidates):
            result.confianza_concursos = "revisar"
            result.confianza_processos = "revisar"
        # Honest reporting: if the official site only served an anti-bot JS
        # challenge, this is a block, not a real miss. Flag it so the reviewer
        # does not waste time hunting for an index that exists behind the wall.
        antibot_block = home.is_antibot

    # --- Collect extra valid URLs (others Tier 3 could have picked) ---
    # Only emitted for buckets a human will actually review: when the chosen URL
    # is confirmado we trust it, so extras would just be unvalidated keyword
    # matches (detail pages, news) cluttering the output. They are kept for
    # revisar / not-found buckets, where alternates help the reviewer pick.
    need_extras_c = result.confianza_concursos != "confirmado"
    need_extras_p = result.confianza_processos != "confirmado"
    if need_extras_c or need_extras_p:
        chosen_urls = {result.url_concursos, result.url_processos_seletivos}
        concurso_kw = {"concurso", "concursos"}
        pss_kw = {"seletivo", "seletivos", "pss", "selecao"}
        extras_c: list[str] = []
        extras_p: list[str] = []
        for c in all_candidates:
            if not c.fetchable or c.url in chosen_urls:
                continue
            text_lower = (c.menu_text or "").lower() + " " + c.url.lower()
            if c.page and c.page.title:
                text_lower += " " + c.page.title.lower()
            if need_extras_c and any(k in text_lower for k in concurso_kw):
                extras_c.append(c.url)
            if need_extras_p and any(k in text_lower for k in pss_kw):
                extras_p.append(c.url)
        if need_extras_c:
            result.urls_extras_concursos = " | ".join(extras_c[:5])
        if need_extras_p:
            result.urls_extras_processos = " | ".join(extras_p[:5])

    notes_parts = []
    total = len(all_candidates)
    fetchable = len([c for c in all_candidates if c.fetchable])
    if total > 0:
        notes_parts.append(f"{total} candidates ({fetchable} fetchable)")
    if not result.url_concursos and not result.url_processos_seletivos:
        notes_parts.append("no valid index page found")
    if antibot_block:
        notes_parts.append("bloqueo_antibot: site responde challenge JS (indice no accesible)")
    result.notes = "; ".join(notes_parts)

    for bucket, selected in result.selected_resources.items():
        result.selected_evidence[bucket] = BucketCandidateEvidence(
            bucket=bucket, candidate=selected.candidate,
            snapshot=selected.candidate.evidence_snapshot,
        )

    return result


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
OUTPUT_FIELDS = [
    "uf", "municipio", "site_base",
    "url_concursos", "confianza_concursos",
    "url_processos_seletivos", "confianza_processos",
    "urls_extras_concursos", "urls_extras_processos",
    "tier_concursos", "tier_processos",
    "method", "razao", "notes", "checked_at",
]


def _read_existing_rows(path: Path) -> dict[str, dict]:
    """Read an existing output CSV into a {norm(municipio): row} map."""
    rows: dict[str, dict] = {}
    if not path.exists():
        return rows
    try:
        with path.open("r", encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                muni = (row.get("municipio") or "").strip()
                if muni:
                    rows[norm(muni)] = row
    except Exception as e:
        print(f"Could not read existing CSV for append: {e}", flush=True)
    return rows


def _preserve_confirmed_buckets(result: MunicipioResult,
                                existing_row: dict | None) -> list[str]:
    """Keep already-confirmed bucket fields when --skip-existing retries a row."""
    if not existing_row:
        return []
    preserved: list[str] = []
    bucket_fields = [
        ("concursos", "url_concursos", "confianza_concursos",
         "tier_concursos", "urls_extras_concursos"),
        ("processos", "url_processos_seletivos", "confianza_processos",
         "tier_processos", "urls_extras_processos"),
    ]
    for bucket, url_field, conf_field, tier_field, extras_field in bucket_fields:
        if existing_row.get(conf_field) != "confirmado":
            continue
        setattr(result, url_field, existing_row.get(url_field, ""))
        setattr(result, conf_field, "confirmado")
        setattr(result, tier_field, existing_row.get(tier_field, ""))
        setattr(result, extras_field, existing_row.get(extras_field, ""))
        preserved.append(bucket)
    if preserved:
        if existing_row.get("site_base") and not result.site_base:
            result.site_base = existing_row.get("site_base", "")
        note = f"skip_existing_preserved:{','.join(preserved)}"
        result.notes = f"{result.notes}; {note}" if result.notes else note
    return preserved


def write_results(results: list[MunicipioResult], path: Path,
                  append: bool = False, csv_only: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    now = datetime.now(timezone.utc).isoformat()

    # Start from existing rows when appending, then overlay this run.
    merged: dict[str, dict] = _read_existing_rows(path) if append else {}
    for r in results:
        merged[norm(r.municipio)] = {
            "uf": UF_SIGLA,
            "municipio": r.municipio,
            "site_base": r.site_base,
            "url_concursos": r.url_concursos,
            "confianza_concursos": r.confianza_concursos,
            "url_processos_seletivos": r.url_processos_seletivos,
            "confianza_processos": r.confianza_processos,
            "urls_extras_concursos": r.urls_extras_concursos,
            "urls_extras_processos": r.urls_extras_processos,
            "tier_concursos": r.tier_concursos,
            "tier_processos": r.tier_processos,
            "method": r.method,
            "razao": r.razao,
            "notes": r.notes,
            "checked_at": now,
        }

    ordered = sorted(merged.values(), key=lambda d: norm(d.get("municipio", "")))
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        for row in ordered:
            writer.writerow({k: row.get(k, "") for k in OUTPUT_FIELDS})
    if csv_only:
        return
    print(f"\nCSV written to {path} ({len(ordered)} rows)", flush=True)

    # --- Excel output with proper formatting ---
    xlsx_path = path.with_suffix(".xlsx")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Concursos RS"

        # Colors for confidence levels
        fills = {
            "confirmado": PatternFill("solid", fgColor="C6EFCE"),  # green
            "probable":   PatternFill("solid", fgColor="FFEB9C"),  # yellow
            "revisar":    PatternFill("solid", fgColor="FFC7CE"),  # red/pink
        }
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        link_font = Font(color="0563C1", underline="single", size=10)
        wrap_align = Alignment(wrap_text=True, vertical="top")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # User-friendly column names
        excel_cols = [
            ("Municipio", 22),
            ("Site Base", 30),
            ("URL Concursos", 45),
            ("Confianza C", 14),
            ("URL Processos Seletivos", 45),
            ("Confianza P", 14),
            ("URLs Extras Concursos", 40),
            ("URLs Extras Processos", 40),
            ("Tier C", 8),
            ("Tier P", 8),
            ("Razon IA", 60),
            ("Notas", 35),
            ("Fecha", 22),
        ]
        for col_idx, (name, width) in enumerate(excel_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = width

        for row_idx, row in enumerate(ordered, 2):
            vals = [
                row.get("municipio", ""), row.get("site_base", ""),
                row.get("url_concursos", ""), row.get("confianza_concursos", ""),
                row.get("url_processos_seletivos", ""), row.get("confianza_processos", ""),
                row.get("urls_extras_concursos", ""), row.get("urls_extras_processos", ""),
                row.get("tier_concursos", ""), row.get("tier_processos", ""),
                row.get("razao", ""), row.get("notes", ""), row.get("checked_at", now),
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = wrap_align
                cell.border = thin_border

            # Color-code confidence columns
            for conf_col in (4, 6):
                cell = ws.cell(row=row_idx, column=conf_col)
                if cell.value in fills:
                    cell.fill = fills[cell.value]

            # Make URLs clickable
            for url_col in (2, 3, 5):
                cell = ws.cell(row=row_idx, column=url_col)
                if cell.value and str(cell.value).startswith("http"):
                    cell.font = link_font
                    cell.hyperlink = str(cell.value)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
        print(f"Excel written to {xlsx_path}", flush=True)
    except ImportError:
        print("openpyxl not installed, skipping Excel output", flush=True)
    except Exception as e:
        print(f"Excel write error: {e}", flush=True)


# ---------------------------------------------------------------------------
# Municipality loading
# ---------------------------------------------------------------------------
def load_municipios_from_golden(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [row["municipio"] for row in reader if row.get("municipio")]


def load_municipios_from_tce(session: requests.Session,
                             timeout: int = 30) -> list[str]:
    try:
        resp = session.get(DEFAULT_MUNICIPIOS_URL, timeout=timeout)
        resp.raise_for_status()
        # Sniff the delimiter (the TCE export is comma-separated).
        sample = resp.text[:512]
        delim = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(resp.text.splitlines(), delimiter=delim)
        return sorted(set(
            (row["NOME_MUNICIPIO"] or "").strip().title()
            for row in reader
            if (row.get("UF") or row.get("SIGLA_UF", "")) == "RS"
            and row.get("NOME_MUNICIPIO")
        ))
    except Exception as e:
        print(f"Error loading municipios: {e}", file=sys.stderr)
        return []


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="5-tier cascade for RS municipality resource discovery"
    )
    parser.add_argument("--golden", type=Path,
                        default=PROJECT_ROOT / "authority_first" / "data" / "golden_set_v1.csv",
                        help="Run only on golden set municipalities")
    parser.add_argument("--all", action="store_true",
                        help="Run on all 497 RS municipalities")
    parser.add_argument("--municipio", type=str,
                        help="Run on a single municipality")
    parser.add_argument("--output", type=Path,
                        default=PROJECT_ROOT / "data" / "cascade_output.csv")
    parser.add_argument("--model", type=str,
                        default=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"))
    parser.add_argument("--gemini-free-first", action="store_true",
                        help="Use GEMINI_API_KEY_FREE/GEMINI_FREE_MODEL first, "
                             "falling back to GEMINI_API_KEY on quota/limits")
    parser.add_argument("--no-playwright", action="store_true",
                        help="Skip Tier 4")
    parser.add_argument("--timeout", type=int, default=15)
    parser.add_argument("--limit", type=int, default=0,
                        help="Limit number of municipalities")
    parser.add_argument("--letras", type=str, default="",
                        help="Only process municipalities whose name starts with "
                             "one of these letters (accent-insensitive), e.g. 'ab'")
    parser.add_argument("--append", action="store_true",
                        help="Merge into the existing output CSV instead of "
                             "overwriting it (rows for the same municipality are "
                             "replaced; new ones are appended)")
    parser.add_argument("--skip-existing", action="store_true",
                        help="Skip municipalities whose concursos and processos "
                             "buckets are both already confirmed in the output CSV. "
                             "If only one bucket is confirmed, retry the municipality "
                             "and preserve that confirmed bucket. Combine with --append.")
    parser.add_argument("--grounded-verify", action="store_true",
                        help="Compatibility flag retained; final decisions now use the "
                             "single CandidateRecord adjudicator and never a second "
                             "grounded verification authority")
    args = parser.parse_args()

    if args.gemini_free_first:
        os.environ["GEMINI_FREE_FIRST"] = "1"

    session = make_session()

    if args.municipio:
        municipios = [args.municipio]
    elif args.all:
        municipios = load_municipios_from_tce(session)
    else:
        municipios = load_municipios_from_golden(args.golden)

    if args.letras:
        wanted = {c for c in norm(args.letras) if c.isalnum()}
        municipios = [m for m in municipios if norm(m)[:1] in wanted]

    existing_rows: dict[str, dict] = {}
    if args.skip_existing:
        existing_rows = _read_existing_rows(args.output)
        settled = {
            key for key, row in existing_rows.items()
            if row.get("confianza_concursos") == "confirmado"
            and row.get("confianza_processos") == "confirmado"
        }
        before = len(municipios)
        municipios = [m for m in municipios if norm(m) not in settled]
        skipped = before - len(municipios)
        if skipped:
            print(f"Skipping {skipped} fully-confirmed municipalities "
                  f"(--skip-existing); re-processing {len(municipios)}", flush=True)

    if args.limit > 0:
        municipios = municipios[:args.limit]

    print(f"Processing {len(municipios)} municipalities", flush=True)
    print(f"Model: {args.model}", flush=True)
    print(f"Playwright: {'disabled' if args.no_playwright else 'enabled'}", flush=True)
    print("=" * 60, flush=True)

    results = []
    for i, muni in enumerate(municipios, 1):
        print(f"\n[{i}/{len(municipios)}] {muni}", flush=True)
        try:
            r = process_municipio(
                session, muni, args.model,
                timeout=args.timeout,
                use_playwright=not args.no_playwright,
            )
            preserved = []
            if args.skip_existing:
                preserved = _preserve_confirmed_buckets(
                    r, existing_rows.get(norm(muni)))
            results.append(r)
            if preserved:
                print(f"  preserved confirmed bucket(s): {', '.join(preserved)}",
                      flush=True)
            status = []
            if r.url_concursos:
                status.append(f"C: {r.url_concursos}")
            if r.url_processos_seletivos:
                status.append(f"P: {r.url_processos_seletivos}")
            if not status:
                status.append("nothing found")
            print(f"  result → {'; '.join(status)}", flush=True)
        except Exception as e:
            print(f"  ERROR: {e}", flush=True)
            traceback.print_exc()
            r = MunicipioResult(municipio=muni, notes=f"error: {e}")
            preserved = []
            if args.skip_existing:
                preserved = _preserve_confirmed_buckets(
                    r, existing_rows.get(norm(muni)))
            results.append(r)
            if preserved:
                print(f"  preserved confirmed bucket(s): {', '.join(preserved)}",
                      flush=True)

        # Checkpoint after every municipality (CSV only, fast) so a crash/stop
        # loses nothing: re-running with --skip-existing resumes where it left off.
        try:
            write_results(results, args.output,
                          append=args.append or args.skip_existing, csv_only=True)
        except Exception as e:
            print(f"  checkpoint write failed: {e}", flush=True)

    # --- Compatibility closure for uncertain legacy results ---
    to_verify: list[dict] = []
    verify_index: dict[str, MunicipioResult] = {}
    for r in results:
        for bucket, url, conf in [
            ("concursos", r.url_concursos, r.confianza_concursos),
            ("processos", r.url_processos_seletivos, r.confianza_processos),
        ]:
            if conf == "probable" and url:
                to_verify.append({
                    "municipio": r.municipio, "bucket": bucket,
                    "url": url, "title": "", "preview": "",
                    "site_base": r.site_base,
                })
                verify_index[f"{r.municipio}|{bucket}"] = r

    if to_verify:
        print(f"\n{'='*60}", flush=True)
        print(f"Batch verification: {len(to_verify)} uncertain URLs", flush=True)
        to_verify, verify_index, verdicts = _batch_verify_uncertain_results(
            session, args.model, results, timeout=args.timeout,
            use_playwright=not args.no_playwright,
        )

        confirmed = sum(1 for _, (v, _) in verdicts.items() if v == "confirmado")
        for key, (status, reason) in sorted(verdicts.items()):
            LOGGER.info("batch_bucket %s", json.dumps({
                "key": key, "status": status, "reason": reason,
            }, ensure_ascii=False, sort_keys=True))
        print(f"  Verified: {confirmed}/{len(verdicts)} upgraded to confirmado",
              flush=True)

    if args.grounded_verify:
        LOGGER.info(
            "compatibility_flag %s",
            json.dumps({
                "flag": "grounded-verify",
                "action": "no-op",
                "reason": "FinalDecision usa adjudicador central unico",
            }, ensure_ascii=False, sort_keys=True),
        )

    # --skip-existing implies append: the skipped rows must be preserved.
    write_results(results, args.output, append=args.append or args.skip_existing)

    # --- Summary ---
    found_c = sum(1 for r in results if r.url_concursos)
    found_p = sum(1 for r in results if r.url_processos_seletivos)
    conf_c = sum(1 for r in results if r.confianza_concursos == "confirmado")
    conf_p = sum(1 for r in results if r.confianza_processos == "confirmado")
    prob_c = sum(1 for r in results if r.confianza_concursos == "probable")
    prob_p = sum(1 for r in results if r.confianza_processos == "probable")
    rev_c = sum(1 for r in results if r.confianza_concursos == "revisar")
    rev_p = sum(1 for r in results if r.confianza_processos == "revisar")
    print(f"\nSummary: {found_c}/{len(results)} concursos, "
          f"{found_p}/{len(results)} processos found", flush=True)
    print(f"  Concursos  — confirmado: {conf_c}, probable: {prob_c}, revisar: {rev_c}",
          flush=True)
    print(f"  Processos  — confirmado: {conf_p}, probable: {prob_p}, revisar: {rev_p}",
          flush=True)
    print(f"  Gemini calls — {gemini_post_call_summary()}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
